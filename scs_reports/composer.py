"""Deterministic workbook composer.

COPY VERIFIED MASTER -> MODIFY TARGET CELLS -> REMOVE IRRELEVANT CONTENT
-> ADD NECESSARY INSTANCES SAFELY.

Masters are never written. The planner decides what belongs; this module
decides how Excel is constructed.
"""
from __future__ import annotations

import re
from copy import copy
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema import (
    JobRecord,
    Equipment,
    AirDevice,
    Traverse,
    PhotoEvidence,
    Measurement,
    tested_category_labels,
)
from .planner import ReportPlan, SectionPlan
from .store import JobStore, ReportPaths
from .masters import MASTER_WORKBOOKS, master_path


def _sanitize_stem(value: str) -> str:
    cleaned = re.sub(r"[^\w\- ]+", "", value or "Job").strip()
    return cleaned or "Job"


def output_stem(record: JobRecord) -> str:
    project = _sanitize_stem(record.metadata.project_name)
    number = _sanitize_stem(record.metadata.project_number or "")
    day = (
        record.metadata.test_date.isoformat()
        if record.metadata.test_date
        else datetime.now().strftime("%Y-%m-%d")
    )
    report_kind = (record.metadata.report_type or "").upper()
    kind = "AIRFLOW" if report_kind in ("AIRFLOW_VERIFICATION", "OUTLET_BALANCE") else "TAB"
    return f"{project}_{number}_{kind}_{day}"


def _copy_style(source, target) -> None:
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def _populated_span(ws) -> tuple[int, int]:
    """True used span (row, col) from populated cells, ignoring phantom
    full-column/full-row dimensions that inflate max_row/max_column."""
    cells = getattr(ws, "_cells", None)
    if not cells:
        return ws.max_row, ws.max_column
    last_row = last_col = 0
    for (row, col) in cells:
        if row > last_row:
            last_row = row
        if col > last_col:
            last_col = col
    return last_row, last_col


def copy_sheet(
    source: Worksheet,
    target: Worksheet,
    *,
    row_offset: int = 0,
    max_row: int | None = None,
    max_col: int | None = None,
) -> None:
    span_row, span_col = _populated_span(source)
    if max_row is None:
        max_row = min(span_row, 2000)
    if max_col is None:
        max_col = min(span_col, 400)
    last_row = min(source.max_row, max_row)
    last_col = min(source.max_column, max_col)
    for row in source.iter_rows(min_row=1, max_row=last_row, max_col=last_col):
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            target_cell = target.cell(
                row=cell.row + row_offset, column=cell.column
            )
            if cell.value is not None:
                target_cell.value = cell.value
            _copy_style(cell, target_cell)
    for merged in source.merged_cells.ranges:
        if merged.min_row + row_offset > 0:
            target.merge_cells(
                start_row=merged.min_row + row_offset,
                start_column=merged.min_col,
                end_row=merged.max_row + row_offset,
                end_column=merged.max_col,
            )
    for row_index, dimension in source.row_dimensions.items():
        if dimension.height is not None:
            target.row_dimensions[row_index + row_offset].height = dimension.height
        target.row_dimensions[row_index + row_offset].hidden = dimension.hidden
    for column_index, dimension in source.column_dimensions.items():
        target.column_dimensions[column_index].width = dimension.width
        target.column_dimensions[column_index].hidden = dimension.hidden


def clone_rows(
    source: Worksheet,
    target: Worksheet,
    *,
    first: int,
    last: int,
    target_first: int,
    count: int,
) -> None:
    """Clone a row block `count` times starting at target_first (style copy)."""
    for instance in range(count):
        offset = target_first - first + instance * (last - first + 1)
        for row in source.iter_rows(min_row=first, max_row=last):
            for cell in row:
                target_cell = target.cell(
                    row=cell.row + offset, column=cell.column
                )
                if cell.value is not None:
                    target_cell.value = cell.value
                _copy_style(cell, target_cell)
        for merged in source.merged_cells.ranges:
            if merged.min_row < first or merged.max_row > last:
                continue
            target.merge_cells(
                start_row=merged.min_row + offset,
                start_column=merged.min_col,
                end_row=merged.max_row + offset,
                end_column=merged.max_col,
            )
        for row_index, dimension in source.row_dimensions.items():
            if first <= row_index <= last and dimension.height is not None:
                target.row_dimensions[row_index + offset].height = dimension.height


def clone_columns(
    source: Worksheet,
    target: Worksheet,
    *,
    first: int,
    last: int,
    target_first: int,
    count: int,
    rows: tuple[int, int],
) -> None:
    """Clone a column block `count` times starting at target_first (style copy)."""
    for instance in range(count):
        offset = target_first - first + instance * (last - first + 1)
        for row in source.iter_rows(min_row=rows[0], max_row=rows[1]):
            for cell in row:
                if not (first <= cell.column <= last):
                    continue
                target_cell = target.cell(
                    row=cell.row, column=cell.column + offset
                )
                if cell.value is not None:
                    target_cell.value = cell.value
                _copy_style(cell, target_cell)
        for merged in source.merged_cells.ranges:
            if merged.min_col < first or merged.max_col > last:
                continue
            if merged.min_row < rows[0] or merged.max_row > rows[1]:
                continue
            target.merge_cells(
                start_row=merged.min_row,
                start_column=merged.min_col + offset,
                end_row=merged.max_row,
                end_column=merged.max_col + offset,
            )
        for column_index in range(first, last + 1):
            column = get_column_letter(column_index + offset)
            source_column = get_column_letter(column_index)
            if source.column_dimensions[source_column].width:
                target.column_dimensions[column].width = (
                    source.column_dimensions[source_column].width
                )


def _write_header(
    ws: Worksheet,
    fields: dict[str, str],
    record: JobRecord,
    *,
    job_number: str | None = None,
) -> None:
    metadata = record.metadata
    values = {
        "project": metadata.project_name,
        "site": (
            metadata.site_name
            or metadata.site_address
            or None
        ),
        "contractor": metadata.hiring_contractor,
        "job_number": job_number or metadata.project_number,
        "test_date": metadata.test_date
        if isinstance(metadata.test_date, (datetime, date))
        else (
            datetime.fromisoformat(metadata.test_date)
            if metadata.test_date
            else None
        ),
        "technician": metadata.technician,
    }
    for field, coordinate in fields.items():
        if field in values and values[field] is not None:
            ws[coordinate] = values[field]


class Composer:
    def __init__(
        self,
        paths: ReportPaths | None = None,
        store: JobStore | None = None,
    ) -> None:
        self.paths = paths or ReportPaths().ensure()
        self.store = store or JobStore(self.paths)
        self._masters: dict[str, object] = {}

    # ---------------------------------------------------------------- helpers

    def _load_master(self, filename: str) -> object:
        if filename not in self._masters:
            path = master_path(self.paths.masters, filename)
            self._masters[filename] = load_workbook(
                path, data_only=False, keep_vba=False
            )
        return self._masters[filename]

    def _cache_master(self, master) -> None:
        """Keep a cached master alive across sections (replaces close calls)."""
        return None

    def close_masters(self) -> None:
        """Release cached master workbooks (read-only; never executed VBA)."""
        for workbook in self._masters.values():
            try:
                workbook.close()
            except Exception:
                pass
        self._masters.clear()

    def _add_sheet(self, output, title: str) -> Worksheet:
        if title in output.sheetnames:
            base, counter = title, 2
            while f"{base} ({counter})" in output.sheetnames:
                counter += 1
            title = f"{base} ({counter})"
        return output.create_sheet(title)

    def _compose_header_sheet(self, ws: Worksheet, record: JobRecord) -> None:
        fields = {
            "project": "C6",
            "site": "C7",
            "contractor": "C8",
            "job_number": "I6",
            "test_date": "I7",
            "technician": "I8",
        }
        _write_header(ws, fields, record)

    # ---------------------------------------------------------------- sections

    def _cover(self, output, record: JobRecord) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "02_Cover")
        copy_sheet(master["02_Cover"], ws)
        _write_header(
            ws,
            {
                "report_date": "C8",
                "project": "C10",
                "site": "C12",
                "job_number": "C14",
                "contractor": "C16",
                "technician": "C20",
            },
            record,
            job_number=record.metadata.project_number or record.metadata.job_id,
        )
        ws["C18"] = "TAB Report"
        self._cache_master(master)
        return ws

    def _certification(self, output, record: JobRecord) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "04_Certification")
        copy_sheet(master["04_Certification"], ws)
        self._compose_header_sheet(ws, record)
        ws["E4"] = "STATUS: DRAFT"
        self._cache_master(master)
        return ws

    def _abbreviations(self, output) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "06_Abbreviations")
        copy_sheet(master["06_Abbreviations"], ws)
        self._cache_master(master)
        return ws

    def _executive_summary(self, output, record: JobRecord) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "03_Executive_Summary")
        copy_sheet(master["03_Executive_Summary"], ws)
        _write_header(
            ws,
            {
                "project": "C6",
                "site": "C7",
                "contractor": "C8",
                "job_number": "I6",
                "test_date": "I7",
                "technician": "I8",
            },
            record,
        )
        systems = ", ".join(
            equipment.equipment_id for equipment in record.equipment
        )
        ws["C11"] = (
            "DEFICIENT" if record.findings else "PASS"
        )
        ws["C12"] = systems or "See report sheets"
        ws["E4"] = "STATUS: DRAFT"
        narrative = "\n".join(
            f"{finding.title}: {finding.detail}" for finding in record.findings
        )
        if not narrative and record.field_observations:
            narrative = record.field_observations
        if narrative:
            ws["A30"] = narrative
        self._cache_master(master)
        return ws

    def _scope_summary(self, output, record: JobRecord) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "05_Scope_Summary")
        copy_sheet(master["02_Cover"], ws)
        for coordinate in ("C8", "C10", "C12", "C14", "C16", "C18", "C20", "A24", "A30"):
            ws[coordinate] = None
        _write_header(
            ws,
            {
                "report_date": "C8",
                "project": "C10",
                "site": "C12",
                "job_number": "C14",
                "contractor": "C16",
            },
            record,
        )
        ws["C18"] = "TAB Report - Scope"
        ws["C20"] = record.metadata.technician
        ws["A24"] = "SCOPE: " + (
            record.scope_notes
            or "Field testing of HVAC systems as listed on the report sheets."
        )
        if record.categories_tested:
            for merged in list(ws.merged_cells.ranges):
                if merged.min_row <= 24 <= merged.max_row and merged.min_col == 1:
                    ws.unmerge_cells(str(merged))
                    break
            ws.merge_cells(start_row=24, start_column=1, end_row=25, end_column=8)
            ws["A26"] = "CATEGORIES TESTED: " + ", ".join(
                tested_category_labels(record.categories_tested)
            )
        ws["A30"] = "OBSERVATIONS: " + (
            record.field_observations or record.technician_notes
        )
        self._cache_master(master)
        return ws

    def _rtu_nameplate(self, output, record: JobRecord, rtu_ahus: list[Equipment]) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "07_Equipment_Register")
        copy_sheet(master["07_Equipment_Register"], ws, max_row=60, max_col=14)
        _write_header(
            ws,
            {
                "project": "C6",
                "site": "C7",
                "contractor": "C8",
                "job_number": "G6",
                "test_date": "G7",
                "technician": "G8",
            },
            record,
        )
        for row in range(12, 17):
            if ws.cell(row=row, column=1).value is not None:
                ws.cell(row=row, column=1).value = None
        for index, equipment in enumerate(rtu_ahus):
            row = 12 + index
            ws[f"A{row}"] = equipment.equipment_id
            ws[f"B{row}"] = equipment.equipment_type.value
            ws[f"C{row}"] = equipment.area_served
            ws[f"D{row}"] = equipment.manufacturer
            ws[f"E{row}"] = equipment.model
            ws[f"F{row}"] = equipment.serial
            ws[f"G{row}"] = equipment.design_data.design_tons
            voltage = next(
                (
                    m.value
                    for m in equipment.measurements
                    if m.field in ("voltage", "voltage_phase")
                ),
                None,
            )
            if voltage is not None:
                ws[f"H{row}"] = voltage
            refrigerant = next(
                (m.value for m in equipment.measurements if m.field == "refrigerant"),
                None,
            )
            if refrigerant is not None:
                ws[f"I{row}"] = refrigerant
        remarks = "; ".join(note for note in (e.notes for e in rtu_ahus) if note)
        if remarks:
            ws["A50"] = remarks
        self._cache_master(master)
        return ws

    def _building_pressure(self, output, record: JobRecord, devices: list[AirDevice]) -> Worksheet:
        master = self._load_master("Test and Balance MASTER TEMPLATE 001.xlsx")
        ws = self._add_sheet(output, "Bldg Press Summary")
        copy_sheet(master["Bldg Press Summary"], ws)
        ws["A6"] = f"PROJECT:  {record.metadata.project_name}"
        ws["A7"] = f"PROJECT #  {record.metadata.project_number or ''}"
        if record.metadata.test_date:
            ws["H7"] = record.metadata.test_date
        template = master["Bldg Press Summary"]
        if len(devices) > 24:
            clone_rows(
                template,
                ws,
                first=12,
                last=35,
                target_first=36,
                count=len(devices) - 24,
            )
        oa_design = 0.0
        oa_test = 0.0
        ex_design = 0.0
        ex_test = 0.0
        for index, device in enumerate(devices):
            row = 12 + index
            ws[f"C{row}"] = device.device_id
            is_outside = device.function in ("Outside Air", "Makeup Air")
            is_exhaust = device.function in ("Exhaust Air", "Relief Air")
            design = device.design_cfm
            final = device.final_cfm
            if is_outside:
                ws[f"D{row}"] = design
                ws[f"E{row}"] = final
                oa_design += design or 0
                oa_test += final or 0
            elif is_exhaust:
                ws[f"F{row}"] = design
                ws[f"G{row}"] = final
                ex_design += design or 0
                ex_test += final or 0
            else:
                ws[f"D{row}"] = design
                ws[f"E{row}"] = final
                ws[f"F{row}"] = None
                ws[f"G{row}"] = None
        total_row = 12 + len(devices) if len(devices) <= 24 else 36 + (len(devices) - 24)
        ws[f"C{total_row}"] = "Total"
        ws[f"D{total_row}"] = oa_design or None
        ws[f"E{total_row}"] = oa_test or None
        ws[f"F{total_row}"] = ex_design or None
        ws[f"G{total_row}"] = ex_test or None
        pressure = next(
            (
                r.value
                for r in record.environmental_readings
                if r.field == "building_pressure"
            ),
            None,
        )
        if pressure is not None:
            ws["D39"] = pressure
            ws["E39"] = "NEGATIVE" if pressure < 0 else "POSITIVE"
        self._cache_master(master)
        return ws

    def _air_distribution(
        self, output, record: JobRecord, devices: list[AirDevice]
    ) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "20_Air_Distribution")
        copy_sheet(master["20_Air_Distribution"], ws)
        _write_header(
            ws,
            {
                "project": "C6",
                "site": "C7",
                "contractor": "C8",
                "job_number": "G6",
                "test_date": "G7",
                "technician": "G8",
            },
            record,
        )
        ws["E4"] = "STATUS: FINAL"
        systems = ", ".join(
            dict.fromkeys(e.equipment_id for e in record.equipment if e.equipment_id)
        )
        if systems:
            ws["A11"] = systems
        areas = ", ".join(
            dict.fromkeys(
                d.area_served for d in devices if d.area_served
            )
        )
        if areas:
            ws["A12"] = areas
        design_total = sum(d.design_cfm or 0 for d in devices)
        if design_total:
            ws["G11"] = design_total
        instruments = ", ".join(
            dict.fromkeys(
                d.measurement_method for d in devices if d.measurement_method
            )
        )
        if instruments:
            ws["G12"] = instruments
        for index, device in enumerate(devices):
            row = 17 + index
            ws[f"A{row}"] = device.area_served
            ws[f"B{row}"] = device.device_id
            ws[f"C{row}"] = device.function
            ws[f"D{row}"] = device.size
            ws[f"G{row}"] = device.design_cfm
            ws[f"H{row}"] = device.avg_velocity_fpm
            ws[f"I{row}"] = device.as_found_cfm
            ws[f"J{row}"] = None
            ws[f"K{row}"] = device.final_cfm
            ws[f"N{row}"] = device.status
        self._cache_master(master)
        return ws

    def _static_pressure(self, output, record: JobRecord) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "19_Static_Pressure_Profile")
        copy_sheet(master["19_Static_Pressure_Profile"], ws)
        _write_header(
            ws,
            {
                "project": "C6",
                "site": "C7",
                "contractor": "C8",
                "job_number": "G6",
                "test_date": "G7",
                "technician": "G8",
            },
            record,
        )
        ws["E4"] = "STATUS: FINAL"
        systems = ", ".join(
            dict.fromkeys(e.equipment_id for e in record.equipment if e.equipment_id)
        )
        if systems:
            ws["A11"] = systems
        design = sum(d.design_cfm or 0 for d in record.air_devices)
        if design:
            ws["G11"] = design
        measured = sum(d.final_cfm or 0 for d in record.air_devices)
        if measured:
            ws["G13"] = measured
        points = {
            m.field.rsplit("_", 1)[-1].upper(): m.value
            for e in record.equipment
            for m in e.measurements
            if re.fullmatch(r"sp_[a-i]", m.field, re.IGNORECASE)
        }
        for index in range(9):
            label = chr(ord("A") + index)
            value = points.get(label)
            if value is not None:
                ws[f"D{17 + index}"] = value
        for r in record.environmental_readings:
            if r.field == "static_pressure" and r.location:
                ws[f"D{17 + ord(r.location.upper()) - ord('A')}"] = r.value
        self._cache_master(master)
        return ws

    def _deficiencies(self, output, record: JobRecord) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "30_Deficiencies")
        copy_sheet(master["30_Deficiencies"], ws)
        _write_header(
            ws,
            {
                "project": "C6",
                "site": "C7",
                "contractor": "C8",
                "job_number": "G6",
                "test_date": "G7",
                "technician": "G8",
            },
            record,
        )
        ws["E4"] = "STATUS: FINAL"
        for index, finding in enumerate(record.findings):
            row = 12 + index
            ws[f"A{row}"] = f"DEF-{index + 1:03d}"
            ws[f"C{row}"] = finding.category
            ws[f"D{row}"] = finding.severity
            ws[f"F{row}"] = (
                f"{finding.title}: {finding.detail}" if finding.detail else finding.title
            )
            if finding.evidence_refs:
                ws[f"G{row}"] = ", ".join(finding.evidence_refs)
            ws[f"L{row}"] = "OPEN"
        self._cache_master(master)
        return ws

    def _instrument_calibration(self, output, record: JobRecord) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "05_Instrument_Calibration")
        copy_sheet(master["05_Instrument_Calibration"], ws)
        _write_header(
            ws,
            {
                "project": "C6",
                "site": "C7",
                "contractor": "C8",
                "job_number": "G6",
                "test_date": "G7",
                "technician": "G8",
            },
            record,
        )
        ws["E4"] = "STATUS: FINAL"
        self._cache_master(master)
        return ws

    def _traverse_summary(self, output, record: JobRecord, traverses: list[Traverse]) -> Worksheet:
        master = self._load_master("Test and Balance MASTER TEMPLATE 001.xlsx")
        ws = self._add_sheet(output, "Duct Traverse Summary SP")
        copy_sheet(master["Duct Traverse Summary SP"], ws)
        ws["A6"] = f"PROJECT:  {record.metadata.project_name}"
        ws["A7"] = f"PROJECT #  {record.metadata.project_number or ''}"
        if record.metadata.test_date:
            ws["I7"] = record.metadata.test_date
        template = master["Duct Traverse Summary SP"]
        if len(traverses) > 2:
            clone_rows(
                template,
                ws,
                first=11,
                last=12,
                target_first=13,
                count=len(traverses) - 2,
            )
        for index, traverse in enumerate(traverses):
            row = 11 + index
            ws[f"A{row}"] = traverse.location or traverse.system_id
            ws[f"B{row}"] = traverse.instrument
            ws[f"C{row}"] = traverse.air_temp or "NA"
            ws[f"D{row}"] = traverse.duct_size
            ws[f"E{row}"] = traverse.area_sqft
            if traverse.design_fpm is not None:
                ws[f"F{row}"] = traverse.design_fpm
                ws[f"G{row}"] = traverse.design_cfm
            ws[f"H{row}"] = None
            ws[f"I{row}"] = f"=H{row}*E{row}"
            if traverse.final_fpm is not None:
                ws[f"J{row}"] = traverse.final_fpm
            ws[f"K{row}"] = f"=J{row}*E{row}"
            ws[f"L{row}"] = traverse.sp or "NA"
        self._cache_master(master)
        return ws

    def _traverse_points(self, output, record: JobRecord, traverse: Traverse) -> Worksheet:
        master = self._load_master("Test and Balance MASTER TEMPLATE 001.xlsx")
        ws = self._add_sheet(output, f"Duct Traverse {traverse.system_id}")
        template = master["Duct Traverse "]
        copy_sheet(template, ws, row_offset=0)
        ws["A8"] = f"SYSTEM: {traverse.system_id}"
        ws["C9"] = traverse.location
        for point in traverse.points:
            label_row = 11 + (ord(point.row_label.upper()) - ord("A"))
            if label_row > 17:
                continue
            ws.cell(row=label_row, column=1).value = point.row_label
            ws.cell(
                row=label_row,
                column=1 + (point.column if point.column is not None else 1),
            ).value = point.fpm
        ws["D21"] = traverse.duct_size
        ws["G21"] = traverse.duct_size
        ws["D22"] = traverse.area_sqft
        ws["G22"] = traverse.area_sqft
        ws["D23"] = traverse.sp or "NA"
        ws["G23"] = traverse.sp or "NA"
        if traverse.design_fpm is not None:
            ws["D24"] = traverse.design_fpm
            ws["D25"] = traverse.design_cfm
        if traverse.final_fpm is not None:
            ws["G24"] = traverse.final_fpm
            ws["G25"] = traverse.final_cfm
        self._cache_master(master)
        return ws

    def _vav_data(self, output, record: JobRecord, vavs: list[Equipment]) -> Worksheet:
        master = self._load_master("Test and Balance MASTER TEMPLATE 001.xlsx")
        ws = self._add_sheet(output, "VAV Data")
        copy_sheet(master["VAV Data"], ws)
        ws["A6"] = f"PROJECT:  {record.metadata.project_name}"
        ws["A7"] = f"PROJECT #  {record.metadata.project_number or ''}"
        if record.metadata.test_date:
            ws["G7"] = record.metadata.test_date
        template = master["VAV Data"]
        if len(vavs) > 33:
            clone_rows(
                template,
                ws,
                first=11,
                last=43,
                target_first=44,
                count=len(vavs) - 33,
            )
        for index, vav in enumerate(vavs):
            row = 11 + index
            values = {m.field: m.value for m in vav.measurements}
            ws[f"A{row}"] = vav.tag or vav.equipment_id
            ws[f"B{row}"] = vav.tag or vav.equipment_id
            ws[f"C{row}"] = values.get("size")
            ws[f"D{row}"] = values.get("correction_factor")
            ws[f"E{row}"] = values.get("design_min")
            ws[f"F{row}"] = values.get("design_max")
            ws[f"G{row}"] = values.get("final_min")
            ws[f"H{row}"] = values.get("final_max")
        notes = "; ".join(note for note in (v.notes for v in vavs) if note)
        if notes:
            ws["A46"] = notes
        self._cache_master(master)
        return ws

    def _fan_test(self, output, record: JobRecord, fan: Equipment) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, f"12_Fan_Blower ({fan.equipment_id})")
        copy_sheet(master["12_Fan_Blower"], ws)
        self._compose_header_sheet(ws, record)
        values = {m.field: m.value for m in fan.measurements}
        ws["C11"] = fan.equipment_id
        ws["I11"] = fan.notes or fan.area_served
        ws["C12"] = fan.manufacturer
        ws["I12"] = fan.model
        ws["I14"] = fan.area_served
        if fan.serial:
            ws["C13"] = f"{fan.model or ''} / {fan.serial}".strip(" /")
        left_rows = {
            "airflow_cfm": 18,
            "fan_rpm": 19,
            "fan_inlet_sp": 20,
            "fan_discharge_sp": 21,
            "tesp": 22,
            "vfd_frequency": 23,
            "damper_position": 24,
            "sound_vibration": 25,
        }
        for field, row in left_rows.items():
            if field in values:
                ws[f"B{row}"] = None
                ws[f"C{row}"] = values[field]
                ws[f"E{row}"] = "PASS"
        right_rows = {
            "voltage": (18, "H"),
            "amps_l1": (19, "H"),
            "amps_l2": (20, "H"),
            "amps_l3": (21, "H"),
            "motor_rpm": (22, "H"),
            "motor_hp_bhp": (23, "H"),
            "power_factor": (24, "H"),
            "rotation": (25, "H"),
        }
        for field, (row, target) in right_rows.items():
            if field in values:
                ws[f"I{row}"] = values[field]
                ws[f"K{row}"] = "PASS"
        if fan.deficiencies:
            ws["A35"] = "; ".join(d.description for d in fan.deficiencies)
        ws["C44"] = record.metadata.technician
        if record.metadata.test_date:
            ws["I44"] = record.metadata.test_date
        self._cache_master(master)
        return ws

    def _vfd_report(self, output, record: JobRecord, vfd: Equipment) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, f"33_VFD_Report ({vfd.equipment_id})")
        copy_sheet(master["33_VFD_Report"], ws)
        self._compose_header_sheet(ws, record)
        values = {m.field: m.value for m in vfd.measurements}
        ws["C11"] = vfd.equipment_id
        ws["I11"] = vfd.area_served
        ws["C12"] = vfd.manufacturer
        ws["I12"] = vfd.tag
        if vfd.model:
            ws["C13"] = f"{vfd.model} / {vfd.serial or ''}".strip(" /")
        left_rows = {
            "input_voltage": 18,
            "output_voltage": 19,
            "drive_hp_kw": 20,
            "drive_output_hz": 21,
            "min_frequency": 22,
            "max_frequency": 23,
            "command_source": 24,
            "motor_speed_pct": 25,
        }
        for field, row in left_rows.items():
            if field in values:
                ws[f"B{row}"] = None
                ws[f"C{row}"] = values[field]
                ws[f"E{row}"] = "PASS"
        right_rows = {
            "motor_voltage": 18,
            "motor_fla": 19,
            "amps_l1": 20,
            "amps_l2": 21,
            "amps_l3": 22,
            "motor_rotation": 25,
        }
        for field, row in right_rows.items():
            if field in values:
                ws[f"H{row}"] = None
                ws[f"I{row}"] = values[field]
                ws[f"K{row}"] = "PASS"
        ws["C44"] = record.metadata.technician
        if record.metadata.test_date:
            ws["I44"] = record.metadata.test_date
        self._cache_master(master)
        return ws

    def _photo_log(self, output, record: JobRecord) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "31_Photo_Log")
        copy_sheet(master["31_Photo_Log"], ws)
        self._compose_header_sheet(ws, record)
        for row_index, photo in enumerate(record.photos, start=12):
            ws.cell(row=row_index, column=1, value=photo.photo_id)
            ws.cell(row=row_index, column=3, value=photo.equipment_association)
            ws.cell(row=row_index, column=4, value=photo.classification.value)
            notes = "; ".join(
                f"{f.get('field', '')}={f.get('value', '')}"
                for f in photo.candidate_facts
            )
            ws.cell(row=row_index, column=5, value=notes or photo.original_filename)
            ws.cell(row=row_index, column=8, value=photo.original_filename)
        ws.print_area = f"A1:J{11 + len(record.photos)}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        self._cache_master(master)
        return ws

    def _remarks(self, output, record: JobRecord) -> Worksheet:
        ws = self._add_sheet(output, "Remarks")
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["C"].width = 90
        ws["A1"] = "SUNSHINE CLIMATE SOLUTIONS LLC"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "REMARKS / FIELD NOTES"
        ws["A2"].font = Font(bold=True, size=12)
        ws["A4"] = "FORM: SCS-RMK-001"
        ws["E4"] = "STATUS: DRAFT"
        ws["A6"] = "PROJECT / CLIENT"
        ws["C6"] = record.metadata.project_name
        ws["G6"] = "JOB #"
        ws["I6"] = record.metadata.project_number or record.metadata.job_id
        ws["A7"] = "SERVICE SITE"
        ws["C7"] = record.metadata.site_name or record.metadata.site_address
        ws["G7"] = "TEST DATE"
        ws["I7"] = record.metadata.test_date
        ws["A8"] = "HIRING CONTRACTOR / PO"
        ws["C8"] = record.metadata.hiring_contractor
        ws["G8"] = "TECHNICIAN"
        ws["I8"] = record.metadata.technician
        row = 10
        if record.categories_tested:
            ws.cell(row=row, column=1, value="CATEGORIES TESTED").font = Font(bold=True)
            ws.cell(row=row, column=3, value=", ".join(tested_category_labels(record.categories_tested)))
            row += 2
        for label, text in (
            ("FIELD OBSERVATIONS", record.field_observations),
            ("KNOWN DEFICIENCIES", record.known_deficiencies),
            ("TECHNICIAN NOTES", record.technician_notes),
        ):
            if not text:
                continue
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=3, value=text)
            row += 2
        return ws

    def _closeout(self, output, record: JobRecord, plan: ReportPlan) -> Worksheet:
        master = self._load_master("Field_Report_Master.xlsm")
        ws = self._add_sheet(output, "32_Final_Closeout")
        copy_sheet(master["32_Final_Closeout"], ws)
        self._compose_header_sheet(ws, record)
        systems = ", ".join(e.equipment_id for e in record.equipment)
        ws["C11"] = "DEFICIENT" if record.findings or record.known_deficiencies else "PASS"
        ws["C12"] = systems or "See report sheets"
        ws["I12"] = record.known_deficiencies or "None observed"
        if record.metadata.test_date:
            ws["I11"] = record.metadata.test_date
        section_types = {s.type for s in plan.sections}
        ws["B17"] = "Yes"
        ws["C17"] = "Yes"
        ws["B18"] = "Yes" if record.equipment else "No"
        ws["C18"] = "Yes" if record.equipment else "N/A"
        ws["B19"] = "Yes"
        ws["C19"] = "Yes" if section_types & {"fan_test", "vfd_report"} else "N/A"
        ws["B20"] = "No"
        ws["C20"] = "No"
        ws["B21"] = "Yes" if record.findings or record.known_deficiencies else "No"
        ws["C21"] = "Yes" if record.findings or record.known_deficiencies else "N/A"
        ws["B22"] = "Yes" if record.photos else "No"
        ws["C22"] = "Complete" if record.photos else "N/A"
        ws["B24"] = "Yes" if section_types & {"traverse_summary", "building_pressure"} else "No"
        ws["C24"] = "Complete" if section_types & {"traverse_summary", "building_pressure"} else "Out of Scope"
        if record.known_deficiencies:
            ws["A33"] = record.known_deficiencies
            ws["B33"] = "See report sheets"
            ws["G33"] = "Review and close"
        self._cache_master(master)
        return ws

    # ---------------------------------------------------------------- assembly

    def compose(
        self,
        record: JobRecord,
        plan: ReportPlan,
        *,
        output_path: Path | None = None,
    ) -> Path:
        from openpyxl import Workbook

        output = Workbook()
        output.remove(output.active)

        rtu_ahus = [
            e for e in record.equipment if e.equipment_type.value in ("RTU", "AHU")
        ]
        fans = [e for e in record.equipment if e.equipment_type.value == "FAN"]
        vfds = [e for e in record.equipment if e.equipment_type.value == "VFD"]
        vavs = [e for e in record.equipment if e.equipment_type.value == "VAV"]

        for section in plan.sections:
            if section.type == "cover":
                self._cover(output, record)
            elif section.type == "certification":
                self._certification(output, record)
            elif section.type == "abbreviations":
                self._abbreviations(output)
            elif section.type == "executive_summary":
                self._executive_summary(output, record)
            elif section.type == "scope_summary":
                self._scope_summary(output, record)
            elif section.type == "rtu_nameplate":
                self._rtu_nameplate(output, record, rtu_ahus)
            elif section.type == "building_pressure":
                self._building_pressure(output, record, record.air_devices)
            elif section.type == "air_distribution":
                self._air_distribution(output, record, record.air_devices)
            elif section.type == "static_pressure":
                self._static_pressure(output, record)
            elif section.type == "deficiencies":
                self._deficiencies(output, record)
            elif section.type == "equipment_register":
                self._rtu_nameplate(output, record, record.equipment)
            elif section.type == "instrument_calibration":
                self._instrument_calibration(output, record)
            elif section.type == "traverse_summary":
                self._traverse_summary(output, record, record.traverses)
            elif section.type == "traverse_points":
                traverse = next(
                    (
                        t
                        for t in record.traverses
                        if t.system_id == section.system_id
                    ),
                    None,
                )
                if traverse is not None:
                    self._traverse_points(output, record, traverse)
            elif section.type == "vav_data":
                self._vav_data(output, record, vavs)
            elif section.type == "fan_test":
                fan = next(
                    (e for e in fans if e.equipment_id == section.equipment_id),
                    None,
                )
                if fan is not None:
                    self._fan_test(output, record, fan)
            elif section.type == "vfd_report":
                vfd = next(
                    (e for e in vfds if e.equipment_id == section.equipment_id),
                    None,
                )
                if vfd is not None:
                    self._vfd_report(output, record, vfd)
            elif section.type == "photo_log":
                self._photo_log(output, record)
            elif section.type == "remarks":
                self._remarks(output, record)
            elif section.type == "closeout":
                self._closeout(output, record, plan)

        if output_path is None:
            stem = output_stem(record)
            output_path = self.store.next_output_version(record.metadata.job_id, stem)

        output.save(output_path)
        output.close()
        return Path(output_path)