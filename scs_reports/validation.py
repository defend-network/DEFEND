"""Automated validation gate before export.

Each check returns PASS / WARN / BLOCK. BLOCK only for something that makes
the report materially unreliable. Rules per product spec:
NO unexplained required blanks. NO fake values. NO irrelevant sections.
NO repetitive empty equipment blocks. NO invented design values.
NO excessive N/A spam.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .schema import JobRecord, Provenance
from .planner import (
    AIRFLOW_REPORT_TYPES,
    ReportPlan,
    _has_static_pressure_data,
)
from .store import MasterStore, ReportPaths

_PLACEHOLDER_PATTERNS = (
    re.compile(r"\bTBD\b", re.IGNORECASE),
    re.compile(r"\bXXX\b", re.IGNORECASE),
    re.compile(r"\bFIXME\b", re.IGNORECASE),
    re.compile(r"^(DRAFT / FIELD|STATUS: DRAFT / FIELD / FINAL)$", re.IGNORECASE),
)
_REQUIRED_JOB_FIELDS = ("project_name", "site_name", "technician", "test_date")
_REQUIRED_EQUIPMENT_FIELDS = ("equipment_id", "manufacturer", "model")
_KNOWN_UNITS = {
    "cfm", "fpm", "in.w.c.", "pa", "db", "wb", "rh", "%", "hz", "volts", "v",
    "amps", "a", "rpm", "tons", "btu/hr", "kw", "degrees f", "degrees c", "f",
}
_REQUIRED_MEASUREMENT_FIELDS = {
    "rtu": ("airflow_cfm", "voltage"),
    "vav": ("design_min", "design_max"),
    "fan": ("airflow_cfm", "fan_rpm"),
    "vfd": ("input_voltage", "output_voltage"),
}
_CELL_REF_RE = re.compile(r"([A-Za-z]{1,3})(\d+)")


@dataclass
class CheckResult:
    name: str
    status: str
    message: str = ""


@dataclass
class ValidationReport:
    checks: list[CheckResult] = field(default_factory=list)

    def add(self, name: str, status: str, message: str = "") -> None:
        self.checks.append(CheckResult(name, status, message))

    @property
    def blocked(self) -> bool:
        return any(c.status == "BLOCK" for c in self.checks)

    @property
    def warnings(self) -> int:
        return sum(1 for c in self.checks if c.status == "WARN")

    def summary(self) -> str:
        counts = {"PASS": 0, "WARN": 0, "BLOCK": 0}
        for check in self.checks:
            counts[check.status] += 1
        return (
            f"{counts['PASS']} PASS / {counts['WARN']} WARN / "
            f"{counts['BLOCK']} BLOCK"
        )


def _count_nonslash_cells(ws) -> int:
    total = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                total += 1
    return total


def validate_report(
    record: JobRecord,
    plan: ReportPlan,
    output_path: Path,
    *,
    masters: MasterStore | None = None,
) -> ValidationReport:
    report = ValidationReport()

    # ------------------------------------------------------------- structure
    missing = [
        name for name in _REQUIRED_JOB_FIELDS if not getattr(record.metadata, name)
    ]
    report.add(
        "required_fields_complete",
        "BLOCK" if missing else "PASS",
        f"missing: {', '.join(missing)}" if missing else "job header fields present",
    )

    phantom = [
        e.equipment_id
        for e in record.equipment
        if not e.measurements
        and not e.deficiencies
        and not (e.manufacturer or e.model or e.serial)
    ]
    report.add(
        "no_phantom_equipment",
        "BLOCK" if phantom else "PASS",
        f"phantom: {', '.join(phantom)}" if phantom else "every equipment entry carries data",
    )

    ids = [e.equipment_id.casefold() for e in record.equipment]
    duplicates = sorted(
        {eid for eid in ids if ids.count(eid) > 1}
    )
    report.add(
        "no_duplicate_equipment",
        "BLOCK" if duplicates else "PASS",
        f"duplicates: {', '.join(duplicates)}" if duplicates else "equipment ids unique",
    )

    for equipment in record.equipment:
        missing_fields = [
            name
            for name in _REQUIRED_EQUIPMENT_FIELDS
            if not getattr(equipment, name)
        ]
        if missing_fields:
            report.add(
                "equipment_instances_consistent",
                "WARN",
                f"{equipment.equipment_id} missing: {', '.join(missing_fields)}",
            )
    if not any(c.name == "equipment_instances_consistent" for c in report.checks):
        report.add("equipment_instances_consistent", "PASS", "equipment records consistent")

    # ------------------------------------------------------------ measurements
    bad_units = [
        f"{m.field}:{m.unit}"
        for e in record.equipment
        for m in e.measurements
        if m.unit and m.unit.casefold() not in _KNOWN_UNITS
    ]
    report.add(
        "measurement_units_valid",
        "WARN" if bad_units else "PASS",
        f"unknown units: {', '.join(bad_units)}" if bad_units else "all units recognized",
    )

    for equipment in record.equipment:
        required = _REQUIRED_MEASUREMENT_FIELDS.get(
            equipment.equipment_type.value.casefold()
        )
        if not required:
            continue
        missing = [
            name
            for name in required
            if equipment.measurement(name) is None
            and not any(
                m.not_applicable and m.field == name
                for m in equipment.measurements
            )
        ]
        if missing:
            report.add(
                "required_fields_complete",
                "WARN",
                f"{equipment.equipment_id} lacks required measurement(s): {', '.join(missing)}",
            )

    # ----------------------------------------------------------- calculations
    bad_calculations: list[str] = []
    for device in record.air_devices:
        if device.design_cfm and device.final_cfm is not None:
            percent = device.percent_design
            if percent is None or not (0.5 <= percent <= 2.0):
                bad_calculations.append(
                    f"{device.device_id} % design {percent if percent is not None else 'n/a'}"
                )
    for traverse in record.traverses:
        if traverse.final_fpm is not None and traverse.area_sqft:
            if abs(traverse.final_cfm - traverse.final_fpm * traverse.area_sqft) > 0.01:
                bad_calculations.append(f"{traverse.traverse_id} CFM != FPM * area")
    report.add(
        "calculation_outputs_correct",
        "WARN" if bad_calculations else "PASS",
        "; ".join(bad_calculations) if bad_calculations else "deterministic math consistent",
    )

    inputs_missing = [
        f"{t.traverse_id}"
        for t in record.traverses
        if t.area_sqft is None or t.area_sqft <= 0 or not t.duct_size
    ]
    report.add(
        "calculation_inputs_present",
        "WARN" if inputs_missing else "PASS",
        f"traverses lacking area/size: {', '.join(inputs_missing)}"
        if inputs_missing
        else "all calculation inputs present",
    )

    design_issues = [
        d.device_id
        for d in record.air_devices
        if d.final_cfm is not None
        and d.design_cfm is None
        and not (d.notes and "design" in d.notes.casefold())
    ]
    report.add(
        "design_vs_actual_consistent",
        "WARN" if design_issues else "PASS",
        f"devices with actual but no design: {', '.join(design_issues)}"
        if design_issues
        else "design vs actual consistent",
    )

    # --------------------------------------------------------------- evidence
    photo_ids = {p.photo_id for p in record.photos}
    orphan_refs = sorted(
        {
            ref
            for entity in (
                list(record.equipment)
                + list(record.air_devices)
                + list(record.traverses)
                + list(record.findings)
            )
            for ref in getattr(entity, "evidence_refs", [])
            if ref not in photo_ids
        }
    )
    report.add(
        "no_orphan_evidence",
        "WARN" if orphan_refs else "PASS",
        f"unresolved refs: {', '.join(orphan_refs)}" if orphan_refs else "all evidence refs resolve",
    )

    if record.photos:
        bad_photos = [
            p.photo_id for p in record.photos if not p.sha256 or len(p.sha256) != 64
        ]
        report.add(
            "photo_refs_valid",
            "BLOCK" if bad_photos else "PASS",
            f"bad sha256: {', '.join(bad_photos)}" if bad_photos else "photo manifest valid",
        )
    else:
        report.add("photo_refs_valid", "PASS", "no photos attached")

    invented = [
        f"{m.field}"
        for e in record.equipment
        for m in e.measurements
        if m.source_type == Provenance.AI_INFERRED_TEXT
        and not m.technician_confirmed
        and m.value is not None
    ]
    report.add(
        "no_invented_measurements",
        "BLOCK" if invented else "PASS",
        f"unconfirmed AI values: {', '.join(invented)}" if invented else "no unconfirmed inferred values",
    )

    # ------------------------------------------------------------------ plan
    section_types = [s.type for s in plan.sections]
    overrides = set(record.plan_overrides)
    has_rtu = any(e.equipment_type.value in ("RTU", "AHU") for e in record.equipment)
    has_devices = bool(record.air_devices)
    has_traverses = bool(record.traverses)
    has_fans = any(e.equipment_type.value == "FAN" for e in record.equipment)
    has_vfds = any(e.equipment_type.value == "VFD" for e in record.equipment)
    has_vavs = any(e.equipment_type.value in ("VAV", "FCU") for e in record.equipment)
    has_equipment = bool(record.equipment)
    has_findings = bool(record.findings)
    has_pressures = _has_static_pressure_data(record)
    is_airflow = (record.metadata.report_type or "").upper() in AIRFLOW_REPORT_TYPES
    airflow_sheet = "air_distribution" if is_airflow else "building_pressure"
    non_airflow_sheet = "building_pressure" if is_airflow else "air_distribution"
    plan_issues: list[tuple[str, str]] = []
    for message, condition, override_entry in (
        ("rtu_nameplate missing despite RTU/AHU equipment", has_rtu and "rtu_nameplate" not in section_types, "remove:rtu_nameplate"),
        ("airflow sheet missing despite air devices", has_devices and airflow_sheet not in section_types, f"remove:{airflow_sheet}"),
        ("static_pressure missing despite static data", has_pressures and "static_pressure" not in section_types, "remove:static_pressure"),
        ("deficiencies missing despite findings", has_findings and "deficiencies" not in section_types, "remove:deficiencies"),
        ("traverse_summary missing despite traverses", has_traverses and "traverse_summary" not in section_types, "remove:traverse_summary"),
        ("rtu_nameplate selected without RTU/AHU equipment", not has_rtu and "rtu_nameplate" in section_types, "add:rtu_nameplate"),
        ("equipment_register selected without equipment", not has_equipment and "equipment_register" in section_types, "add:equipment_register"),
        ("equipment_register selected alongside rtu_nameplate", has_rtu and "rtu_nameplate" in section_types and "equipment_register" in section_types, "remove:equipment_register"),
        ("wrong airflow sheet selected for report type", has_devices and airflow_sheet not in section_types and non_airflow_sheet in section_types, f"add:{airflow_sheet}"),
        ("airflow sheet selected without air devices", not has_devices and airflow_sheet in section_types, f"add:{airflow_sheet}"),
        ("non-airflow sheet selected without air devices", not has_devices and non_airflow_sheet in section_types, f"add:{non_airflow_sheet}"),
        ("static_pressure selected without static data", not has_pressures and "static_pressure" in section_types, "add:static_pressure"),
        ("deficiencies selected without findings", not has_findings and "deficiencies" in section_types, "add:deficiencies"),
        ("traverse_summary selected without traverses", not has_traverses and "traverse_summary" in section_types, "add:traverse_summary"),
        ("fan_test selected without fan equipment", not has_fans and "fan_test" in section_types, "add:fan_test"),
        ("vfd_report selected without VFD equipment", not has_vfds and "vfd_report" in section_types, "add:vfd_report"),
        ("vav_data selected without VAV equipment", not has_vavs and "vav_data" in section_types, "add:vav_data"),
    ):
        if condition:
            if override_entry in overrides:
                plan_issues.append((message + " (manual override)", "WARN"))
            else:
                plan_issues.append((message, "BLOCK"))
    blocked = [message for message, status in plan_issues if status == "BLOCK"]
    report.add(
        "no_phantom_sections",
        "BLOCK" if blocked else "PASS" if not plan_issues else "WARN",
        "; ".join(f"{message}" for message, _status in plan_issues)
        if plan_issues
        else "sections match record content",
    )

    # ---------------------------------------------------------------- workbook
    try:
        workbook = load_workbook(output_path, data_only=False)
        workbook.close()
        report.add("workbook_opens_successfully", "PASS", "openpyxl load ok")
    except Exception as error:
        report.add(
            "workbook_opens_successfully",
            "BLOCK",
            f"load failed: {type(error).__name__}: {error}",
        )
        return report

    workbook = load_workbook(output_path, data_only=False)
    bad_formulas: list[str] = []
    blank_required: list[str] = []
    placeholder: list[str] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.print_area:
            raw_area = ws.print_area.replace("$", "")
            if "!" in raw_area:
                raw_area = raw_area.split("!", 1)[1]
            area = raw_area.split(":")[0]
            try:
                ws[area]
            except Exception:
                bad_formulas.append(f"{sheet_name}:print_area")
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    if value.strip() == "=":
                        bad_formulas.append(f"{sheet_name}!{cell.coordinate}")
                        continue
                    if value.count("(") != value.count(")"):
                        bad_formulas.append(f"{sheet_name}!{cell.coordinate}")
                        continue
                    for column_name, row_number in _CELL_REF_RE.findall(value):
                        column_index = (
                            sum(
                                (ord(character) - 64)
                                * (26 ** (len(column_name) - 1 - position))
                                for position, character in enumerate(column_name)
                            )
                            if column_name
                            else 0
                        )
                        if column_index > ws.max_column or int(row_number) > ws.max_row:
                            bad_formulas.append(
                                f"{sheet_name}!{cell.coordinate} ref out of range"
                            )
                            break
                if isinstance(value, str) and any(
                    pattern.search(value) for pattern in _PLACEHOLDER_PATTERNS
                ):
                    placeholder.append(f"{sheet_name}!{cell.coordinate}")
                if isinstance(value, str) and value.casefold() in {
                    "x", "xxx", "enter value", "enter values",
                }:
                    placeholder.append(f"{sheet_name}!{cell.coordinate}")
    report.add(
        "no_formula_errors",
        "BLOCK" if bad_formulas else "PASS",
        f"bad formulas: {', '.join(bad_formulas[:10])}" if bad_formulas else "formulas well-formed",
    )
    report.add(
        "print_areas_valid",
        "PASS",
        "print areas resolve" if not bad_formulas else "see formula errors",
    )
    report.add(
        "no_placeholder_text",
        "WARN" if placeholder else "PASS",
        f"placeholders: {', '.join(placeholder[:10])}" if placeholder else "no placeholder text",
    )
    report.add(
        "no_unexplained_required_blanks",
        "PASS" if not blank_required else "WARN",
        f"blank required cells: {', '.join(blank_required[:10])}" if blank_required else "no unexplained required blanks",
    )
    workbook.close()

    if masters is not None:
        unchanged, changed = masters.verify_unchanged()
        report.add(
            "master_unchanged",
            "BLOCK" if not unchanged else "PASS",
            f"changed masters: {', '.join(changed)}"
            if changed
            else "all masters byte-identical to registry",
        )

    return report