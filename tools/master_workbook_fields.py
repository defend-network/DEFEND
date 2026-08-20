"""M2b: field-level scan of the two owner-approved master workbooks.

Extends the structural audit (master_workbook_audit.py) with FIELD CONCEPTS:
for every sheet, walks the label region (first ~32 rows, all columns) and
collects non-empty text cells as candidate field labels, including merged-cell
spans, so the report-driven vision catalog can map canonical fields to real
workbook destinations (sheet + label/cell).

Also scans the filled example reports in C:\\SCS_DATA\\masters\\ so real-world
field usage (Fan Test, Equipment Register, Traverse, etc.) is captured.

Emits C:\\SCS_DATA\\masters\\field_scan.json. Read-only: no write access.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    sys.exit(2)

MASTERS = [
    "Field_Report_Master.xlsm",
    "Test and Balance MASTER TEMPLATE 001.xlsx",
]
EXAMPLES = [
    "SCS-BP-RTU-Data-Only.xlsx",
    "SCS-Gatorade-Report.xlsm",
    "SCS-LakePanasoffkee-Traverse.xlsx",
    "SCS-Roland-VAV.xlsx",
]

_LABEL_RE = re.compile(r"[A-Za-z][A-Za-z0-9 /&()'.-]{1,60}")
_SKIP = {"", None}


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def scan_workbook(path: Path) -> dict:
    out: dict = {"path": str(path), "sheets": []}
    try:
        wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    except Exception as exc:
        out["open_error"] = str(exc)
        return out
    for ws in wb.worksheets:
        rows: list[dict] = []
        merged_map: dict[str, str] = {}
        for rng in getattr(ws, "merged_cells", ()) or ():
            try:
                start = rng.min_row, rng.min_col
                merged_map[f"{rng.min_row}:{rng.min_col}"] = (
                    f"{rng.max_row}:{rng.max_col}")
            except Exception:
                continue
        seen: set[str] = set()
        for row in ws.iter_rows(max_row=min(ws.max_row or 1, 80)):
            for cell in row:
                text = cell_text(cell.value)
                if not text or len(text) > 80:
                    continue
                coord = f"{cell.row}:{cell.column}"
                if coord in seen:
                    continue
                seen.add(coord)
                rows.append({
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                    "label": text,
                    "merged_to": merged_map.get(coord),
                })
        out["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "cells": rows,
        })
    try:
        wb.close()
    except Exception:
        pass
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--masters", default=r"C:\SCS_DATA\masters")
    parser.add_argument("--out", default=r"C:\SCS_DATA\masters\field_scan.json")
    args = parser.parse_args()
    root = Path(args.masters)
    result: dict = {}
    for name in MASTERS + EXAMPLES:
        path = root / name
        if not path.exists():
            print(f"[scan] missing {path}", file=sys.stderr)
            continue
        scan = scan_workbook(path)
        result[name] = scan
        print(f"[scan] {name}: {len(scan.get('sheets', []))} sheets, "
              f"{sum(len(s.get('cells', [])) for s in scan.get('sheets', []))} label cells")
    Path(args.out).write_text(json.dumps(result, indent=1), encoding="utf-8")
    print(f"\n[scan] wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())