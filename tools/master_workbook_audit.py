"""M2: audit the two owner-approved master workbooks.

For each master workbook in C:\\SCS_DATA\\masters\\:
  * ZIP inspection: entry list, VBA presence (xl/vbaProject.bin), external
    links (xl/externalLinks/), shared strings count.
  * openpyxl inspection (keep_vba for .xlsm): sheet names, dimensions, merged
    ranges, named ranges, data validations, print areas, page setup, images,
    charts, formula-bearing cells.

Emits C:\\SCS_DATA\\masters\\audit.json and a human-readable summary. The
audit is read-only: no workbook is ever opened with write access.
"""

from __future__ import annotations

import argparse
import json
import sys
import zipfile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    sys.exit(2)

FIELD_REPORT = "Field_Report_Master.xlsm"
TAB_TEMPLATE = "Test and Balance MASTER TEMPLATE 001.xlsx"
FORMULA_PREFIXES = ("=", "=" + "{", "=" + "=")
LINK_NAMES = {"Table", "DefinedName", "Print_Area", "Print_Titles"}


def audit_zip(path: Path) -> dict:
    result = {"entries": [], "has_vba": False, "external_links": [], "shared_strings": 0}
    try:
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
            result["entries"] = sorted(names)
            result["has_vba"] = "xl/vbaProject.bin" in names
            result["external_links"] = [
                n for n in names if n.startswith("xl/externalLinks/")
            ]
            if "xl/sharedStrings.xml" in names:
                try:
                    content = zf.read("xl/sharedStrings.xml").decode("utf-8", "replace")
                    result["shared_strings"] = content.count("<si>")
                except Exception:
                    result["shared_strings"] = -1
    except Exception as exc:
        result["zip_error"] = str(exc)
    return result


def audit_workbook(path: Path) -> dict:
    out: dict = {
        "path": str(path),
        "bytes": path.stat().st_size,
        "sha256": None,
        "zip": {},
        "sheets": [],
        "named_ranges": [],
        "data_validations": 0,
        "images": 0,
        "charts": 0,
        "formula_cells": 0,
        "print_areas": [],
    }
    out["sha256"] = _sha256(path)
    out["zip"] = audit_zip(path)
    try:
        wb = load_workbook(path, read_only=True, data_only=False, keep_vba=True)
    except Exception as exc:
        out["open_error"] = str(exc)
        return out

    for ws in wb.worksheets:
        sheet = {
            "title": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged": len(getattr(ws, "merged_cells", ()) or []),
            "print_area": str(getattr(ws, "print_area", None) or ""),
            "orientation": getattr(getattr(ws, "page_setup", None), "orientation", None),
            "paper_size": getattr(getattr(ws, "page_setup", None), "paperSize", None),
            "validations": len(getattr(ws, "data_validations", ()) or []),
        }
        if getattr(ws, "images", None):
            sheet["image_count"] = len(ws._images) if hasattr(ws, "_images") else len(ws.images)
        if getattr(ws, "_charts", None):
            sheet["chart_count"] = len(ws._charts)
        formula_cells = 0
        try:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells += 1
        except Exception:
            pass
        sheet["formula_cells"] = formula_cells
        out["formula_cells"] += formula_cells
        out["sheets"].append(sheet)

    for name in getattr(wb, "defined_names", {}):
        out["named_ranges"].append(
            {
                "name": name,
                "kind": "range" if name not in LINK_NAMES else "internal",
            }
        )
    if hasattr(wb, "worksheets") and hasattr(wb, "images"):
        pass
    wb.close()
    return out


def _sha256(path: Path) -> str:
    import hashlib

    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest().upper()


def summarize(audit: dict) -> None:
    for key, entry in audit.items():
        name = Path(entry["path"]).name
        print(f"=== {name} ({entry['bytes']:,} bytes, sha256={entry['sha256']})")
        if "open_error" in entry:
            print(f"  OPEN ERROR: {entry['open_error']}")
            continue
        print(f"  vba={entry['zip']['has_vba']} shared_strings={entry['zip']['shared_strings']} "
              f"external_links={len(entry['zip']['external_links'])}")
        print(f"  zip entries: {len(entry['zip']['entries'])}")
        for sheet in entry["sheets"]:
            print(
                f"  sheet '{sheet['title']}': {sheet['max_row']}r x {sheet['max_col']}c "
                f"merged={sheet['merged']} validations={sheet['validations']} "
                f"print_area='{sheet['print_area']}' orient={sheet['orientation']} "
                f"paper={sheet['paper_size']} formulas={sheet['formula_cells']}"
                f"{' images=' + str(sheet.get('image_count')) if 'image_count' in sheet else ''}"
                f"{' charts=' + str(sheet.get('chart_count')) if 'chart_count' in sheet else ''}"
            )
        print(f"  named ranges: {len(entry['named_ranges'])} "
              f"({', '.join(r['name'] for r in entry['named_ranges'][:20])})")
        print(f"  total formula cells: {entry['formula_cells']}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--masters", default=r"C:\SCS_DATA\masters")
    parser.add_argument("--out", default=r"C:\SCS_DATA\masters\audit.json")
    args = parser.parse_args()
    masters_dir = Path(args.masters)
    targets = [masters_dir / FIELD_REPORT, masters_dir / TAB_TEMPLATE]
    missing = [t for t in targets if not t.exists()]
    if missing:
        print(f"[audit] missing masters: {[str(m) for m in missing]}", file=sys.stderr)
        return 2
    audit = {FIELD_REPORT: audit_workbook(targets[0]), TAB_TEMPLATE: audit_workbook(targets[1])}
    Path(args.out).write_text(json.dumps(audit, indent=2), encoding="utf-8")
    summarize(audit)
    print(f"\n[audit] wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())