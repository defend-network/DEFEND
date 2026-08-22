"""SCS Blueprint Intelligence V1 tests.

Covers the P35 required list: immutable upload, page indexing, sheet
number/title extraction, schedule detection + parsing, device CFM/size/
instance extraction, room association, design totals, plan-vs-schedule
conflict, design != measured provenance, pre-engineered Air Distribution,
field checklist, ready-to-leave, revision separation, low-confidence guard,
source provenance, cache reuse, no customer PDF committed.
"""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fixtures.blueprint_ground_truth import (
    EXPECTED_DEVICES,
    EXPECTED_EQUIPMENT,
    EXPECTED_SHEETS,
    EXPECTED_SUPPLY_TOTALS,
)
from fixtures.make_blueprint import build_blueprint
from scs_reports.completeness import ready_to_leave
from scs_reports.composer import Composer
from scs_reports.plans import (
    cached_basis,
    index_pdf,
    run_document,
    sha256_of,
)
from scs_reports.planner import plan_for
from scs_reports.preengineer import (
    build_preengineered_record,
    field_test_plan,
)
from scs_reports.schema import (
    JobMetadata,
    JobRecord,
)
from scs_reports.store import JobStore, MasterStore, ReportPaths
from scs_reports.validation import validate_report
from tools.blueprint_benchmark import evaluate as benchmark_evaluate

MASTER_SOURCE = Path(r"C:\SCS_DATA\masters")


@pytest.fixture(scope="module")
def blueprint_pdf(tmp_path_factory) -> Path:
    path = tmp_path_factory.mktemp("blueprint") / "blueprint_fixture.pdf"
    build_blueprint(path)
    return path


@pytest.fixture(scope="module")
def blueprint_pdf_conflict(tmp_path_factory) -> Path:
    path = tmp_path_factory.mktemp("blueprint") / "blueprint_conflict.pdf"
    build_blueprint(path, supply_overrides={"RTU-5": 1999})
    return path


@pytest.fixture(scope="module")
def basis(blueprint_pdf):
    return run_document(blueprint_pdf)


@pytest.fixture
def paths(tmp_path):
    return ReportPaths(tmp_path).ensure()


@pytest.fixture
def masters(paths):
    store = MasterStore(paths)
    store.install_masters(MASTER_SOURCE)
    return store


def _monday_record() -> JobRecord:
    return JobRecord(JobMetadata(
        job_id="monday_bp",
        project_name="Workout Studio Airflow",
        project_number="A-1234",
        site_name="Studio A & B",
        test_date=date(2026, 8, 24),
        technician="AT",
        report_type="AIRFLOW_VERIFICATION",
    ))


# ---------------------------------------------------------------------------
# PDF ingestion / immutability
# ---------------------------------------------------------------------------


def test_pdf_upload_immutable(blueprint_pdf, tmp_path):
    digest_before = sha256_of(blueprint_pdf)
    target = tmp_path / "copy.pdf"
    target.write_bytes(blueprint_pdf.read_bytes())
    assert sha256_of(target) == digest_before  # copy is byte-identical
    assert sha256_of(blueprint_pdf) == digest_before  # source untouched
    doc = index_pdf(target)
    assert doc.sha256 == digest_before


def test_page_indexing(blueprint_pdf):
    doc = index_pdf(blueprint_pdf)
    assert len(doc.pages) == 7
    for page in doc.pages:
        assert page.page_number >= 1
        assert page.words or page.text
        assert page.width > 0 and page.height > 0


# ---------------------------------------------------------------------------
# Sheet classification
# ---------------------------------------------------------------------------


def test_sheet_number_and_title_extraction(blueprint_pdf):
    doc = index_pdf(blueprint_pdf)
    by_page = {p.page_number: p.sheet_number for p in doc.pages}
    for page, (sheet, _type) in EXPECTED_SHEETS.items():
        assert by_page.get(page) == sheet, f"page {page} sheet number"


def test_sheet_classification(blueprint_pdf):
    doc = index_pdf(blueprint_pdf)
    by_page = {p.page_number: p.page_type for p in doc.pages}
    for page, (sheet, ptype) in EXPECTED_SHEETS.items():
        assert by_page.get(page) == ptype, f"page {page} ({sheet}) classified {by_page.get(page)}"


def test_mechanical_sheets_selected(blueprint_pdf):
    doc = index_pdf(blueprint_pdf)
    mechanical = [p for p in doc.pages if p.page_type not in ("ELECTRICAL", "COVER", "UNKNOWN")]
    assert len(mechanical) == 5
    assert all(p.page_type in (
        "MECHANICAL_GENERAL_NOTES", "AIR_DEVICE_SCHEDULE", "EQUIPMENT_SCHEDULE",
        "MECHANICAL_PLAN") for p in mechanical)


# ---------------------------------------------------------------------------
# Schedule extraction
# ---------------------------------------------------------------------------


def test_air_device_schedule_parsed(basis):
    schedule_types = basis["schedule_types"]
    assert "SD-1" in schedule_types
    assert schedule_types["SD-1"]["design_cfm"] == "200"
    assert "SD-2" in schedule_types
    assert schedule_types["SD-2"]["design_cfm"] == "180"
    assert schedule_types["SD-1"]["size"] == "10x10"


def test_equipment_schedule_parsed(basis):
    equipment = {e["tag"]: e for e in basis["equipment"]}
    assert "RTU-5" in equipment and "RTU-6" in equipment
    assert equipment["RTU-5"]["supply_cfm"] == 1180.0
    assert equipment["RTU-5"]["manufacturer"] == "GREENHECK"
    assert equipment["RTU-6"]["supply_cfm"] == 1240.0


# ---------------------------------------------------------------------------
# Device / CFM / size / room extraction
# ---------------------------------------------------------------------------


def test_device_instances_extracted(basis):
    ids = {d["device_id"] for d in basis["instances"]}
    assert len(ids) == len(EXPECTED_DEVICES)
    assert set(EXPECTED_DEVICES) == ids


def test_device_cfm_extracted(basis):
    actual = {d["device_id"]: d for d in basis["instances"]}
    for tag, attrs in EXPECTED_DEVICES.items():
        assert actual[tag]["design_cfm"] == attrs["cfm"], f"{tag} CFM"
        src = actual[tag]["source"]
        assert src.get("sheet") and src.get("page") and src.get("bbox")
        assert src.get("extraction_method") in ("PLAN_CALLOUT", "SCHEDULE_MAPPING")


def test_device_size_extracted(basis):
    actual = {d["device_id"]: d for d in basis["instances"]}
    for tag, attrs in EXPECTED_DEVICES.items():
        assert actual[tag]["size"] == attrs["size"], f"{tag} size"


def test_room_association(basis):
    actual = {d["device_id"]: d for d in basis["instances"]}
    for tag, attrs in EXPECTED_DEVICES.items():
        assert (actual[tag].get("room") or "").upper() == attrs["room"].upper()


def test_design_totals(basis):
    totals = {t["scope"]: t for t in basis["design_totals"] if t["function"] == "SUPPLY"}
    for room, attrs in EXPECTED_SUPPLY_TOTALS.items():
        assert totals[room]["design_total_cfm"] == attrs["cfm"]
        assert totals[room]["device_count"] == attrs["count"]


def test_plan_vs_schedule_conflict_detected(blueprint_pdf_conflict):
    basis = run_document(blueprint_pdf_conflict)
    assert basis["conflicts"], "expected a DESIGN_DOCUMENT_CONFLICT"
    assert any(c["kind"] == "DESIGN_DOCUMENT_CONFLICT" for c in basis["conflicts"])


def test_no_conflict_for_consistent_plans(basis):
    assert basis["conflicts"] == []


# ---------------------------------------------------------------------------
# Pre-engineering
# ---------------------------------------------------------------------------


def test_preengineered_record_design_not_measured(basis):
    record = build_preengineered_record(_monday_record(), basis)
    assert record.metadata.status == "PRE_ENGINEERED"
    for device in record.air_devices:
        assert device.design_cfm is not None
        assert device.as_found_cfm is None
        assert device.final_cfm is None
        assert device.design_source  # provenance preserved
    assert {e.equipment_id for e in record.equipment} == {"RTU-5", "RTU-6"}


def test_design_never_written_into_measured(basis):
    record = build_preengineered_record(_monday_record(), basis)
    from scs_reports.nl_parse import merge_capture, parse_measurements
    before = {d.device_id: d.design_cfm for d in record.air_devices}
    for c in parse_measurements("Studio A SA-1 161 as found; 183 final."):
        merge_capture(record, c)
    sa1 = next(d for d in record.air_devices if d.device_id == "SA-1")
    assert sa1.design_cfm == before["SA-1"]  # unchanged
    assert sa1.as_found_cfm == 161.0
    assert sa1.final_cfm == 183.0


def test_field_checklist_generated(basis):
    record = build_preengineered_record(_monday_record(), basis)
    plan = field_test_plan(record)
    assert len(plan) == len(EXPECTED_DEVICES)
    sa1 = next(p for p in plan if p["device"] == "SA-1")
    assert sa1["room"] == "WORKOUT STUDIO A"
    assert sa1["design_cfm"] == 180.0
    assert sa1["status"] == "NOT MEASURED"


def test_preengineered_air_distribution(paths, masters, basis):
    record = build_preengineered_record(_monday_record(), basis)
    store = JobStore(paths)
    store.create(record)
    plan = plan_for(record)
    assert "air_distribution" in [s.type for s in plan.sections]
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    wb = load_workbook(output)
    ad = wb["20_Air_Distribution"]
    sa1_row = 17
    assert ad.cell(row=sa1_row, column=1).value == "WORKOUT STUDIO A"
    assert ad.cell(row=sa1_row, column=2).value == "SA-1"
    assert ad.cell(row=sa1_row, column=7).value == 180.0       # design CFM
    assert ad.cell(row=sa1_row, column=9).value is None         # as-found blank
    assert ad.cell(row=sa1_row, column=11).value is None        # final blank
    assert ad.cell(row=sa1_row, column=4).value == "8x8"        # size
    wb.close()
    composer.close_masters()


def test_ready_to_leave_catches_unmeasured_planned(basis):
    record = build_preengineered_record(_monday_record(), basis)
    report = ready_to_leave(record)
    assert report["ready"] is False
    assert any("planned_devices_measured" in item for item in report["MISSING_BEFORE_LEAVING"])
    missing_finals = [i for i in report["MISSING_BEFORE_LEAVING"] if ":final_cfm" in i]
    assert len(missing_finals) == len(EXPECTED_DEVICES)
    # once measured, the specific device is no longer flagged
    record.air_devices[0].final_cfm = 183.0
    record.air_devices[0].as_found_cfm = 161.0
    assert f"{record.air_devices[0].device_id}:final_cfm" not in [
        i for i in ready_to_leave(record)["MISSING_BEFORE_LEAVING"]]


def test_design_questions_suppressed_when_plans_present(basis):
    record = build_preengineered_record(_monday_record(), basis)
    report = ready_to_leave(record)
    assert not any("design airflow" in q.lower() for q in report["questions"])


# ---------------------------------------------------------------------------
# Provenance / confidence / revision / cache
# ---------------------------------------------------------------------------


def test_low_confidence_not_silently_autofilled():
    basis = {
        "equipment": [],
        "instances": [
            {"device_id": "SA-99", "room": None, "design_cfm": None, "size": None,
             "schedule_type": None, "source": {"sheet": "M9.9", "page": 9,
                                               "bbox": (0, 0, 1, 1),
                                               "extraction_method": "NATIVE_TEXT"},
             "confidence": "LOW"},
        ],
        "rooms": [], "design_totals": [], "conflicts": [],
    }
    record = build_preengineered_record(_monday_record(), basis)
    device = next(d for d in record.air_devices if d.device_id == "SA-99")
    assert device.design_cfm is None  # never invented from low confidence


def test_source_provenance_for_every_plan_cfm(basis):
    for device in basis["instances"]:
        if device.get("design_cfm") is None:
            continue
        source = device["source"]
        assert source.get("sheet"), "missing sheet provenance"
        assert source.get("page"), "missing page provenance"
        assert source.get("bbox"), "missing bbox provenance"
        assert source.get("extraction_method"), "missing method provenance"


def test_revision_separation(blueprint_pdf, tmp_path):
    import shutil
    rev1 = tmp_path / "rev1.pdf"
    rev2 = tmp_path / "rev2.pdf"
    shutil.copy2(blueprint_pdf, rev1)
    shutil.copy2(blueprint_pdf, rev2)
    doc1 = index_pdf(rev1, revision="REV 1")
    doc2 = index_pdf(rev2, revision="REV 2")
    assert doc1.revision == "REV 1"
    assert doc2.revision == "REV 2"
    # design basis is versioned by document identity, not silently merged
    b1 = cached_basis(rev1, tmp_path / "cache")
    b2 = cached_basis(rev2, tmp_path / "cache")
    assert b1["_cache_hit"] is False
    assert b2["_cache_hit"] is True  # same sha256 reused for identical content


def test_cached_extraction_reused(blueprint_pdf, tmp_path):
    cache_dir = tmp_path / "cache"
    first = cached_basis(blueprint_pdf, cache_dir)
    second = cached_basis(blueprint_pdf, cache_dir)
    assert first["_cache_hit"] is False
    assert second["_cache_hit"] is True
    assert second["document"]["sha256"] == first["document"]["sha256"]


def test_blueprint_benchmark_metrics(basis):
    ground_truth = {
        "EXPECTED_SHEETS": EXPECTED_SHEETS,
        "EXPECTED_EQUIPMENT": EXPECTED_EQUIPMENT,
        "EXPECTED_DEVICES": EXPECTED_DEVICES,
        "EXPECTED_SUPPLY_TOTALS": EXPECTED_SUPPLY_TOTALS,
    }
    metrics = benchmark_evaluate(basis, ground_truth)
    assert metrics["CFM_EXACT_MATCH"] == 1.0
    assert metrics["SIZE_EXACT_MATCH"] == 1.0
    assert metrics["DEVICE_TAG_EXACT_MATCH"] == 1.0
    assert metrics["ROOM_ASSOCIATION_ACCURACY"] == 1.0
    assert metrics["DESIGN_TOTAL_EXACTNESS"] == 1.0
    assert metrics["FALSE_FACT_RATE"] == 0.0


def test_no_customer_pdf_committed():
    from subprocess import run, PIPE
    repo = Path(__file__).resolve().parents[1]
    result = run(["git", "ls-files", "tests", "scs_reports", "tools"],
                 capture_output=True, text=True, cwd=repo)
    tracked = result.stdout.splitlines()
    pdfs = [f for f in tracked if f.lower().endswith(".pdf")]
    assert pdfs == [], f"customer/synthetic PDFs committed to git: {pdfs}"
