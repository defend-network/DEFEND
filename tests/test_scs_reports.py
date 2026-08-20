from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest

from scs_reports.schema import (
    AirDevice,
    Contractor,
    Equipment,
    EquipmentType,
    Finding,
    JobMetadata,
    JobRecord,
    Measurement,
    PhotoEvidence,
    Provenance,
    Traverse,
    TraversePoint,
)
from scs_reports.store import (
    ContractorStore,
    JobStore,
    MasterStore,
    PhotoIngest,
    ReportPaths,
)
from scs_reports.planner import plan_for
from scs_reports.composer import Composer
from scs_reports.validation import validate_report

MASTER_SOURCE = Path(r"C:\SCS_DATA\masters")


@pytest.fixture
def paths(tmp_path):
    return ReportPaths(tmp_path).ensure()


@pytest.fixture
def masters(paths):
    store = MasterStore(paths)
    store.install_masters(MASTER_SOURCE)
    return store


def smoke_record(job_id="smoke_test") -> JobRecord:
    metadata = JobMetadata(
        job_id=job_id,
        project_name="Smoke Test Warehouse",
        project_number="777777",
        site_name="Smoke Site",
        site_address="1 Test Lane",
        test_date=date(2026, 8, 18),
        technician="Aaron Thomas",
        hiring_contractor="Remedy Heating and Cooling",
    )
    record = JobRecord(metadata=metadata)
    rtu = Equipment(
        equipment_id="RTU-1",
        equipment_type=EquipmentType.RTU,
        tag="RTU-1",
        manufacturer="Carrier",
        model="50TC-E08",
        serial="1320P93940",
        area_served="Conditioned gym",
    )
    rtu.measurements = [
        Measurement("voltage", 208, "volts", Provenance.TECH_ENTERED, technician_confirmed=True),
        Measurement("refrigerant", "R-410A", "", Provenance.TECH_ENTERED, technician_confirmed=True),
    ]
    vav = Equipment(
        equipment_id="VAV-1",
        equipment_type=EquipmentType.VAV,
        tag="2-18",
        manufacturer="Titus",
    )
    vav.measurements = [
        Measurement("design_min", 30, "cfm", Provenance.TECH_ENTERED, technician_confirmed=True),
        Measurement("design_max", 220, "cfm", Provenance.TECH_ENTERED, technician_confirmed=True),
    ]
    record.equipment = [rtu, vav]
    record.air_devices = [
        AirDevice("OA-1", "Outside Air", "Gym", 2400, 0, 0, "Velocity Grid", "N/A", 0, "FAIL", "damper closed"),
        AirDevice("EF-1", "Exhaust Air", "Men's", 1445, 1445, 1445, "Velocity Grid", "25.5x18.25", 556, "MEASURED", "ok"),
    ]
    record.traverses = [
        Traverse(
            traverse_id="TRV-1", system_id="OAU-2", location="Catwalk",
            duct_size="18X16", area_sqft=2.0, design_fpm=1200, final_fpm=1108,
            points=[TraversePoint("A", 1344, 1), TraversePoint("B", 1399, 1)],
        )
    ]
    record.findings = [
        Finding("OA damper closed", "RTU-1 OA damper fully closed; zero outdoor airflow.")
    ]
    record.technician_notes = "Synthetic validation data."
    return record


def test_report_paths_layout(paths):
    assert paths.masters.name == "masters"
    assert paths.contractors.name == "contractors"
    assert paths.jobs.name == "jobs"


def test_contractor_add_list_duplicate(paths):
    store = ContractorStore(paths)
    assert store.load() == []
    store.add(Contractor("Remedy Heating and Cooling", contact="Aaron"))
    assert store.find("Remedy Heating and Cooling").contact == "Aaron"
    with pytest.raises(ValueError, match="already exists"):
        store.add(Contractor("remedy heating and cooling"))
    assert len(store.load()) == 1


def test_job_round_trip_persists(paths):
    store = JobStore(paths)
    record = smoke_record()
    store.create(record)
    loaded = store.load("smoke_test")
    assert loaded.metadata.project_name == "Smoke Test Warehouse"
    assert loaded.metadata.test_date == date(2026, 8, 18)
    assert len(loaded.equipment) == 2
    assert loaded.air_devices[1].final_cfm == 1445
    assert loaded.traverses[0].final_cfm == 2216.0
    with pytest.raises(ValueError, match="already exists"):
        store.create(smoke_record())


def test_photo_ingest_sha256_manifest(paths, tmp_path):
    source = tmp_path / "photos"
    source.mkdir()
    for index in range(2):
        (source / f"img_{index}.txt").write_text(f"evidence {index}", encoding="utf-8")
    entries = PhotoIngest(paths).ingest("smoke_test", sorted(source.iterdir()))
    assert len(entries) == 2
    assert entries[0].photo_id == "PHOTO-001"
    assert entries[1].original_filename == "img_1.txt"
    assert len(entries[0].sha256) == 64
    stored = paths.job_subdir("smoke_test", "originals")
    assert (stored / "img_0.txt").exists()
    for entry in entries:
        assert (stored / entry.original_filename).read_text(encoding="utf-8").startswith("evidence")


def test_master_immutability_detects_tampering(paths, masters):
    unchanged, _changed = masters.verify_unchanged()
    assert unchanged is True
    target = paths.masters / "SCS-Roland-VAV.xlsx"
    with target.open("ab") as handle:
        handle.write(b"tampered")
    unchanged, changed = masters.verify_unchanged()
    assert unchanged is False
    assert "SCS-Roland-VAV.xlsx" in changed


def test_planner_selects_only_content_driven_sections():
    record = smoke_record()
    plan = plan_for(record)
    types = [s.type for s in plan.sections]
    assert "rtu_nameplate" in types
    assert "building_pressure" in types
    assert "traverse_summary" in types
    assert "traverse_points" in types
    assert "vav_data" in types
    assert "photo_log" not in types
    assert "fan_test" not in types
    assert "vfd_report" not in types
    assert types[0] == "cover"
    assert types[-1] == "closeout"


def test_planner_no_phantom_sections_for_empty_job():
    record = JobRecord(
        metadata=JobMetadata(job_id="x", project_name="p", site_name="s", technician="t", test_date=date.today())
    )
    plan = plan_for(record)
    types = [s.type for s in plan.sections]
    assert types == ["cover", "certification", "closeout"]


def test_compose_creates_versioned_output_and_correct_sheets(paths, masters):
    store = JobStore(paths)
    record = smoke_record()
    store.create(record)
    record.photos = [
        PhotoEvidence(photo_id="PHOTO-001", original_filename="a.jpg", sha256="0" * 64)
    ]
    store.save(record)
    plan = plan_for(record)
    composer = Composer(paths, store)
    first = composer.compose(record, plan)
    second = composer.compose(record, plan)
    assert first.name == "Smoke Test Warehouse_777777_TAB_2026-08-18.xlsx"
    assert second.name == "Smoke Test Warehouse_777777_TAB_2026-08-18_v02.xlsx"
    from openpyxl import load_workbook

    wb = load_workbook(second)
    assert wb.sheetnames == [
        "02_Cover", "04_Certification", "06_Abbreviations",
        "03_Executive_Summary", "07_Equipment_Register",
        "Bldg Press Summary", "Duct Traverse Summary SP",
        "Duct Traverse OAU-2", "VAV Data", "31_Photo_Log", "Remarks", "32_Final_Closeout",
    ]
    assert wb["02_Cover"]["C10"].value == "Smoke Test Warehouse"
    assert wb["07_Equipment_Register"]["A12"].value == "RTU-1"
    assert wb["07_Equipment_Register"]["B12"].value == "RTU"
    assert wb["Bldg Press Summary"]["E12"].value == 0
    assert wb["Bldg Press Summary"]["F13"].value == 1445
    assert wb["Duct Traverse Summary SP"]["K11"].value == "=J11*E11"
    assert wb["31_Photo_Log"]["A12"].value == "PHOTO-001"
    wb.close()


def test_validation_gates_block_and_warn(paths, masters):
    store = JobStore(paths)
    record = smoke_record()
    store.create(record)
    plan = plan_for(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    assert not report.blocked
    names = {c.name: c.status for c in report.checks}
    assert names["master_unchanged"] == "PASS"
    assert names["no_formula_errors"] == "PASS"
    assert names["workbook_opens_successfully"] == "PASS"
    assert names["no_phantom_sections"] == "PASS"
    assert names["no_phantom_equipment"] == "PASS"


def test_validation_blocks_phantom_equipment(paths, masters):
    record = smoke_record()
    record.equipment.append(Equipment(equipment_id="RTU-9", equipment_type=EquipmentType.RTU, tag="RTU-9"))
    plan = plan_for(record)
    output = path = None
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    block = next(c for c in report.checks if c.name == "no_phantom_equipment")
    assert block.status == "BLOCK"
    assert "RTU-9" in block.message


def test_validation_blocks_unconfirmed_inferred_values(paths, masters):
    record = smoke_record()
    record.equipment[0].measurements.append(
        Measurement(
            "airflow_cfm", 2500, "cfm",
            Provenance.AI_INFERRED_TEXT, source_ref="PHOTO-001",
            confidence=0.55, technician_confirmed=False,
        )
    )
    plan = plan_for(record)
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    block = next(c for c in report.checks if c.name == "no_invented_measurements")
    assert block.status == "BLOCK"


def test_validation_blocks_duplicate_equipment(paths, masters):
    record = smoke_record()
    record.equipment.append(record.equipment[0])
    plan = plan_for(record)
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    block = next(c for c in report.checks if c.name == "no_duplicate_equipment")
    assert block.status == "BLOCK"


def test_validation_blocks_missing_required_fields(paths, masters):
    record = smoke_record()
    record.metadata.technician = ""
    plan = plan_for(record)
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    block = next(c for c in report.checks if c.name == "required_fields_complete")
    assert block.status == "BLOCK"


def test_validation_blocks_phantom_section_in_plan(paths, masters):
    record = smoke_record()
    plan = plan_for(record)
    plan.sections.append(type("S", (), {"type": "fan_test"})())
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    block = next(c for c in report.checks if c.name == "no_phantom_sections")
    assert block.status == "BLOCK"
    assert "without fan equipment" in block.message


def test_validation_warns_on_orphan_evidence(paths, masters):
    record = smoke_record()
    record.findings[0].evidence_refs = ["PHOTO-999"]
    plan = plan_for(record)
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    check = next(c for c in report.checks if c.name == "no_orphan_evidence")
    assert check.status == "WARN"
    assert "PHOTO-999" in check.message


def test_output_never_overwrites_previous_version(paths, masters):
    store = JobStore(paths)
    record = smoke_record()
    store.create(record)
    plan = plan_for(record)
    composer = Composer(paths, store)
    first = composer.compose(record, plan)
    second = composer.compose(record, plan)
    third = composer.compose(record, plan)
    assert first.exists() and second.exists() and third.exists()
    assert first.read_bytes() != second.read_bytes()


def test_plan_overrides_remove_drops_section(paths, masters):
    record = smoke_record()
    record.plan_overrides = ["remove:building_pressure"]
    plan = plan_for(record)
    types = [s.type for s in plan.sections]
    assert "building_pressure" not in types
    assert "rtu_nameplate" in types
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    check = next(c for c in report.checks if c.name == "no_phantom_sections")
    assert check.status == "WARN"
    assert "manual override" in check.message
    assert not report.blocked


def test_plan_overrides_add_allows_empty_section(paths, masters):
    record = JobRecord(
        metadata=JobMetadata(job_id="plain", project_name="Plain", site_name="S", technician="T", test_date=date.today())
    )
    record.plan_overrides = ["add:building_pressure"]
    plan = plan_for(record)
    types = [s.type for s in plan.sections]
    assert "building_pressure" in types
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    check = next(c for c in report.checks if c.name == "no_phantom_sections")
    assert check.status == "WARN"
    assert "manual override" in check.message


def test_not_applicable_measurement_satisfies_required_fields(paths, masters):
    record = smoke_record()
    record.equipment[0].measurements.append(
        Measurement("airflow_cfm", None, "cfm", Provenance.TECH_ENTERED, not_applicable=True)
    )
    plan = plan_for(record)
    store = JobStore(paths)
    store.create(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    report = validate_report(record, plan, output, masters=masters)
    warns = [
        c.message
        for c in report.checks
        if c.name == "required_fields_complete" and c.status == "WARN"
    ]
    assert all("airflow_cfm" not in message for message in warns)


def test_plan_overrides_round_trip_through_store(paths):
    store = JobStore(paths)
    record = smoke_record()
    record.plan_overrides = ["remove:photo_log", "add:fan_test"]
    store.create(record)
    loaded = store.load("smoke_test")
    assert loaded.plan_overrides == ["remove:photo_log", "add:fan_test"]


def test_categories_tested_round_trip(paths):
    store = JobStore(paths)
    record = smoke_record()
    record.categories_tested = ["RTU", "VAV", "Building pressure"]
    store.create(record)
    loaded = store.load("smoke_test")
    assert loaded.categories_tested == ["RTU", "VAV", "Building pressure"]


def test_categories_tested_composed_human_readable(paths, masters):
    store = JobStore(paths)
    record = smoke_record()
    record.scope_notes = "Field testing of listed HVAC systems."
    record.categories_tested = ["RTU", "AHU", "Traverse", "Building pressure"]
    store.create(record)
    plan = plan_for(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    from openpyxl import load_workbook

    wb = load_workbook(output)
    assert wb["05_Scope_Summary"]["A26"].value == (
        "CATEGORIES TESTED: RTU / AHU, Duct Traverse, Building Pressure"
    )
    assert wb["Remarks"]["A10"].value == "CATEGORIES TESTED"
    assert wb["Remarks"]["C10"].value == "RTU / AHU, Duct Traverse, Building Pressure"
    wb.close()


def test_categories_tested_unknown_values_pass_through(paths, masters):
    store = JobStore(paths)
    record = smoke_record()
    record.categories_tested = ["RTU", "Custom category"]
    store.create(record)
    plan = plan_for(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    from openpyxl import load_workbook

    wb = load_workbook(output)
    assert wb["Remarks"]["A10"].value == "CATEGORIES TESTED"
    assert wb["Remarks"]["C10"].value == "RTU / AHU, Custom category"
    wb.close()


def test_categories_tested_empty_omits_line(paths, masters):
    store = JobStore(paths)
    record = smoke_record()
    record.categories_tested = []
    store.create(record)
    plan = plan_for(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    from openpyxl import load_workbook

    wb = load_workbook(output)
    values = [
        cell.value
        for row in wb["Remarks"].iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "CATEGORIES TESTED" not in values
    wb.close()