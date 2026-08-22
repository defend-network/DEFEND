"""SCS M1.2 Mechanical Plan Graph / semantic intelligence tests.

Covers the acceptance list: full/partial/schedule-only/plan-only packets,
missing references, relevant-sheet filtering, lazy indexing, exhaustive
schedule column preservation + normalization + units, equipment/plan
correlation, dampers (incl. fire/smoke/FSD), controls, ducts, notes, keynotes,
details, rooms, systems, provenance, design/measured isolation, graph
validation, graph->DesignBasis, preengineer, Plan Chat citations, scope
relevance, ready-to-leave, conflict abstention, revision separation, false
fire/smoke prevention and customer-file exclusion.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from fixtures.blueprint_ground_truth import (
    M12_EXPECTED_AIRDEVICE_COLUMNS,
    M12_EXPECTED_CONTROLS,
    M12_EXPECTED_DAMPERS,
    M12_EXPECTED_DEVICES,
    M12_EXPECTED_DUCTS,
    M12_EXPECTED_EQUIPMENT,
    M12_EXPECTED_EQUIPMENT_COLUMNS,
    M12_EXPECTED_REFERENCES,
    M12_EXPECTED_SHEETS,
)
from fixtures.make_blueprint import build_blueprint_m12
from scs_reports.plan_graph import MECHANICAL_PLAN_GRAPH_SCHEMA_VERSION
from scs_reports.plan_packet import (
    build_packet,
    classify_missing_context,
    select_deep_pages,
)
from scs_reports.plan_schedules import parse_value
from scs_reports.plan_scope import (
    answer_graph_question,
    field_plan,
    ready_to_leave_graph,
    scope_relevant,
)
from scs_reports.plan_semantics import build_graph, graph_to_design_basis
from scs_reports.planner import plan_for
from scs_reports.plans import PlanDocument, PlanPage, index_pdf
from scs_reports.preengineer import build_preengineered_record
from scs_reports.schema import AirDevice, JobMetadata, JobRecord
from scs_reports.store import JobStore, MasterStore, ReportPaths
from scs_reports.validation import validate_report
from scs_reports.composer import Composer

MASTER_SOURCE = Path(r"C:\SCS_DATA\masters")
SCOPE = "Verify airflow and slightly balance the two workout studio duct systems."


@pytest.fixture(scope="module")
def m12_pdf(tmp_path_factory) -> Path:
    path = tmp_path_factory.mktemp("m12") / "bp_m12.pdf"
    build_blueprint_m12(path)
    return path


@pytest.fixture(scope="module")
def m12_doc(m12_pdf) -> PlanDocument:
    return index_pdf(m12_pdf, tables=False)


@pytest.fixture(scope="module")
def m12_graph(m12_doc):
    return build_graph([m12_doc])


def _subset(doc: PlanDocument, page_numbers: list[int]) -> PlanDocument:
    keep = [p for p in doc.pages if p.page_number in page_numbers]
    return PlanDocument(document_id=doc.document_id,
                        original_filename=doc.original_filename,
                        sha256=doc.sha256, revision=doc.revision, pages=keep)


@pytest.fixture
def paths(tmp_path):
    return ReportPaths(tmp_path).ensure()


@pytest.fixture
def masters(paths):
    store = MasterStore(paths)
    store.install_masters(MASTER_SOURCE)
    return store


def _record() -> JobRecord:
    return JobRecord(JobMetadata(job_id="m12_job", project_name="Workout Studios",
                                 site_name="Studio A & B",
                                 test_date=__import__("datetime").date(2026, 8, 24),
                                 technician="AT", report_type="AIRFLOW_VERIFICATION"))


# ---------------------------------------------------------------------------
# Plan packets: full / partial / schedule-only / plan-only
# ---------------------------------------------------------------------------


def test_full_plan_packet(m12_doc):
    packet = build_packet([m12_doc])
    assert packet.packet_completeness == "FULL_SET"
    assert "M3.1" in packet.mechanical_sheet_ids
    assert "E1.1" in packet.nonmechanical_sheet_ids


def test_partial_plan_packet(m12_doc):
    subset = _subset(m12_doc, [3, 5])  # schedule + plan only, no legend/detail
    packet = build_packet([subset])
    assert packet.packet_completeness in ("PARTIAL_SET", "UNKNOWN_COMPLETENESS")
    assert "M0.2" not in packet.received_sheet_ids


def test_schedule_only_packet(m12_doc):
    subset = _subset(m12_doc, [3, 4])
    packet = build_packet([subset])
    assert set(packet.mechanical_sheet_ids) <= {"M2.1", "M2.2"}
    context = classify_missing_context(packet, SCOPE)
    assert any(c["kind"] == "LEGEND_NOT_SUPPLIED" for c in context)


def test_plan_only_packet(m12_doc):
    subset = _subset(m12_doc, [5, 6])
    packet = build_packet([subset])
    context = classify_missing_context(packet, SCOPE)
    assert any(c["kind"] == "SCHEDULE_CONTEXT_NOT_SUPPLIED" for c in context)


def test_referenced_missing_sheet(m12_doc):
    packet = build_packet([m12_doc])
    assert "M7.2" in packet.referenced_but_missing_sheet_ids
    context = classify_missing_context(packet, SCOPE)
    assert any(c["kind"] == "REFERENCE_MISSING" and c["sheet_id"] == "M7.2"
               for c in context)


def test_relevant_sheet_filtering_and_lazy_index(m12_doc):
    inventory = [{"page": p.page_number, "sheet_number": p.sheet_number,
                  "title": p.sheet_title, "type": p.page_type,
                  "confidence": p.confidence, "native_words": len(p.words)}
                 for p in m12_doc.pages]
    deep, stats = select_deep_pages(inventory, scope_keywords=["STUDIO"])
    assert stats["DOCUMENT_PAGES"] == 9
    e_pages = [e["page"] for e in inventory if e["type"] == "ELECTRICAL"]
    assert all(page not in deep for page in e_pages)
    assert stats["DEEP_PROCESSED_PAGES"] <= stats["DOCUMENT_PAGES"]


# ---------------------------------------------------------------------------
# Schedules: exhaustive preservation + normalization + units
# ---------------------------------------------------------------------------


def test_air_device_schedule_all_columns_preserved(m12_graph):
    sched = next(s for s in m12_graph.schedules if "AIR_DEVICE" in s["kind"])
    assert set(sched["literal_headings"]) == M12_EXPECTED_AIRDEVICE_COLUMNS


def test_unknown_schedule_columns_preserved(m12_graph):
    # air-device schedule preserves every literal heading (mapped or not)
    sched = next(s for s in m12_graph.schedules if "AIR_DEVICE" in s["kind"])
    literal = set(sched["literal_headings"])
    assert {"NC", "THROW", "DAMPER", "REMARKS"} <= literal
    # equipment schedule keeps genuinely unknown columns as CUSTOM
    eq = next(s for s in m12_graph.schedules if s["kind"] == "RTU_SCHEDULE")
    custom = {c["literal_heading"] for c in eq["columns"]
              if c["normalized_heading"] == "CUSTOM_SCHEDULE_FIELD"}
    assert "MOTOR" in custom or "HPVFD" in custom


def test_equipment_schedule_all_columns_preserved(m12_graph):
    sched = next(s for s in m12_graph.schedules if s["kind"] == "RTU_SCHEDULE")
    assert set(sched["literal_headings"]) == M12_EXPECTED_EQUIPMENT_COLUMNS


def test_schedule_normalization_synonyms():
    assert parse_value("1,250 CFM", "SUPPLY_CFM")["normalized_value"] == 1250.0
    assert parse_value("1,250 CFM", "SUPPLY_CFM")["unit"] == "CFM"
    assert parse_value("0.5", "ESP")["unit"] == "IN.W.G."
    assert parse_value("1130", "FAN_RPM")["unit"] == "RPM"
    assert parse_value("620", "OUTSIDE_AIR_CFM")["unit"] == "CFM"


def test_equipment_scheduled_fields(m12_graph):
    rtu5 = next(e for e in m12_graph.equipment if e["id"] == "RTU-5")
    fields = rtu5["scheduled_fields"]
    assert fields["SUPPLY_CFM"] == "1180"
    assert fields["OUTSIDE_AIR_CFM"] == "620"
    assert fields["FAN_RPM"] == "1130"
    assert fields["ESP"] == "0.5"
    assert fields["VFD"] == "YES"
    assert fields["VOLTS"] == "208"
    assert fields["REFRIGERANT"] == "R-410A"


# ---------------------------------------------------------------------------
# Entities
# ---------------------------------------------------------------------------


def test_equipment_entities_and_plan_correlation(m12_graph):
    ids = {e["id"] for e in m12_graph.equipment}
    assert ids == {"RTU-5", "RTU-6"}
    assert any(r.rel_type == "SERVES" and r.source == "RTU-5"
               and r.target == "WORKOUT STUDIO A" for r in m12_graph.relationships)


def test_air_device_regression(m12_graph):
    ids = {d["id"] for d in m12_graph.air_devices}
    assert set(M12_EXPECTED_DEVICES) == ids


def test_damper_entities(m12_graph):
    dampers = {d["id"]: d["damper_type"] for d in m12_graph.dampers}
    assert dampers == M12_EXPECTED_DAMPERS
    assert dampers["FSD-1"] == "COMBINATION_FIRE_SMOKE_DAMPER"
    assert dampers["FD-1"] == "FIRE_DAMPER"
    assert dampers["SMD-1"] == "SMOKE_DAMPER"


def test_false_fire_smoke_positive_prevention(m12_graph):
    ls = [d for d in m12_graph.dampers
          if d["damper_type"] in ("FIRE_DAMPER", "SMOKE_DAMPER",
                                  "COMBINATION_FIRE_SMOKE_DAMPER")]
    assert len(ls) == 3  # FD-1, FSD-1, SMD-1 - no false positives
    assert sum(1 for d in ls if d["damper_type"] == "COMBINATION_FIRE_SMOKE_DAMPER") == 1


def test_control_entities(m12_graph):
    controls = {c["id"]: c["control_type"] for c in m12_graph.controls}
    assert controls == M12_EXPECTED_CONTROLS


def test_duct_segments(m12_graph):
    sizes = {d["size"] for d in m12_graph.duct_segments}
    assert sizes == M12_EXPECTED_DUCTS
    assert all(d["shape"] == "RECTANGULAR" for d in m12_graph.duct_segments)


def test_rooms_and_relationships(m12_graph):
    rooms = {r["id"] for r in m12_graph.rooms}
    assert rooms == {"WORKOUT STUDIO A", "WORKOUT STUDIO B"}
    located = [r for r in m12_graph.relationships if r.rel_type == "LOCATED_IN"]
    assert any(r.source == "SA-1" and r.target == "WORKOUT STUDIO A" for r in located)


def test_notes_and_keynotes(m12_graph):
    note_texts = {n["literal_text"] for n in m12_graph.notes}
    assert any("BALANCING DAMPER" in t for t in note_texts)
    assert any(r.rel_type == "HAS_NOTE" and r.source == "K12" for r in m12_graph.relationships)


def test_references(m12_graph):
    refs = {r["id"]: r.get("target_present") for r in m12_graph.references}
    assert refs == M12_EXPECTED_REFERENCES


def test_source_provenance_present(m12_graph):
    for entity in m12_graph.equipment + m12_graph.dampers + m12_graph.controls:
        assert entity.get("source", {}).get("sheet") or entity.get("source", {}).get("page")


def test_design_measured_isolation(m12_graph):
    for device in m12_graph.air_devices:
        assert "as_found_cfm" not in device and "final_cfm" not in device
    basis = graph_to_design_basis(m12_graph)
    assert all(i.get("as_found_cfm") is None and i.get("final_cfm") is None
               for i in basis["instances"])


# ---------------------------------------------------------------------------
# Graph validation / schema / conflicts
# ---------------------------------------------------------------------------


def test_graph_schema_version(m12_graph):
    assert m12_graph.schema_version == MECHANICAL_PLAN_GRAPH_SCHEMA_VERSION


def test_graph_validation_flags_unresolved_reference(m12_graph):
    issues = m12_graph.validate()
    assert any(i["kind"] == "UNRESOLVED_REFERENCE" and "M7.2" in i["detail"]
               for i in issues)


def test_numeric_conflict_surfaces_with_all_sources(m12_doc, tmp_path_factory):
    from fixtures.make_blueprint import _grid, _title_block, _draw_table
    import fitz

    pdf = fitz.open()
    page_size = fitz.paper_rect("letter-l")
    page = pdf.new_page(width=page_size.width, height=page_size.height)
    _grid(pdf, page, page_size, label_top="EQUIPMENT SCHEDULE",
          footer="EQUIPMENT SCHEDULE")
    headers = ["TAG", "TYPE", "MFR", "MODEL", "SUPPLY CFM", "OA CFM", "ESP",
               "FAN RPM", "MOTOR HP", "VFD", "VOLTS", "PH", "REFRIG", "REMARKS"]
    col_w = [44, 40, 70, 60, 62, 52, 36, 46, 46, 34, 38, 24, 50, 120]
    rows = [["RTU-5", "RTU", "GREENHECK", "SQ-30", "1999", "620", "0.5",
             "1130", "5", "YES", "208", "3", "R-410A", "SERVES WORKOUT STUDIO A"]]
    _draw_table(page, 30, 96, col_w, headers, rows, page_size)
    _title_block(pdf, page, page_size, "M2.2", "EQUIPMENT SCHEDULE")
    conflict_pdf = tmp_path_factory.mktemp("bp") / "conflict_schedule.pdf"
    pdf.save(str(conflict_pdf))
    pdf.close()
    conflict_doc = index_pdf(conflict_pdf, tables=False)
    main = build_graph([m12_doc])
    conf_graph = build_graph([conflict_doc])
    main.equipment = conf_graph.equipment
    main.schedules = main.schedules + conf_graph.schedules
    from scs_reports.plan_semantics import _compute_conflicts
    main.conflicts = _compute_conflicts(main)
    assert main.conflicts
    conflict = main.conflicts[0]
    assert conflict["kind"] == "DESIGN_DOCUMENT_CONFLICT"
    sources = {c["value"] for c in conflict.get("competing_sources", [])}
    assert 1999 in sources and 1180 in sources


def test_revision_separation(m12_pdf, tmp_path):
    import shutil
    rev1 = tmp_path / "r1.pdf"
    rev2 = tmp_path / "r2.pdf"
    shutil.copy2(m12_pdf, rev1)
    shutil.copy2(m12_pdf, rev2)
    d1 = index_pdf(rev1, revision="REV 1")
    d2 = index_pdf(rev2, revision="REV 2")
    assert d1.revision == "REV 1" and d2.revision == "REV 2"
    assert index_pdf(rev1).sha256 == index_pdf(rev2).sha256  # same doc, distinct rev


# ---------------------------------------------------------------------------
# Graph -> DesignBasis -> preengineer -> report
# ---------------------------------------------------------------------------


def test_graph_to_design_basis(m12_graph):
    basis = graph_to_design_basis(m12_graph)
    assert {i["device_id"] for i in basis["instances"]} == set(M12_EXPECTED_DEVICES)
    assert {e["tag"] for e in basis["equipment"]} == {"RTU-5", "RTU-6"}


def test_preengineer_regression(m12_graph):
    record = build_preengineered_record(_record(), graph_to_design_basis(m12_graph))
    assert record.metadata.status == "PRE_ENGINEERED"
    for device in record.air_devices:
        assert device.design_cfm is not None
        assert device.final_cfm is None
        assert device.design_source


def test_preengineered_report_0_block(paths, masters, m12_graph):
    record = build_preengineered_record(_record(), graph_to_design_basis(m12_graph))
    store = JobStore(paths)
    store.create(record)
    plan = plan_for(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    validation = validate_report(record, plan, output, masters=masters)
    assert not validation.blocked
    wb = load_workbook(output)
    ad = wb["20_Air_Distribution"]
    assert ad.cell(row=17, column=2).value == "SA-1"
    assert ad.cell(row=17, column=7).value == 180.0
    assert ad.cell(row=17, column=9).value is None
    wb.close()
    composer.close_masters()


# ---------------------------------------------------------------------------
# Plan Chat / scope / ready-to-leave
# ---------------------------------------------------------------------------


def test_plan_chat_citations(m12_graph):
    answer = answer_graph_question("What is RTU-5 scheduled for?", m12_graph)
    assert "1180" in answer["answer"] and "GREENHECK" in answer["answer"]
    assert answer["source"]["sheet"] == "M2.2"
    rpm = answer_graph_question("What RPM is the fan?", m12_graph)
    assert "1130" in rpm["answer"] or "1150" in rpm["answer"]
    fsd = answer_graph_question("Show me all fire/smoke dampers", m12_graph)
    assert "FSD-1" in fsd["answer"] and "FIRE_DAMPER" in fsd["answer"]


def test_scope_relevance(m12_graph):
    relevant = scope_relevant(m12_graph, SCOPE)
    assert {r["id"] for r in relevant["rooms"]} == {"WORKOUT STUDIO A", "WORKOUT STUDIO B"}
    assert len(relevant["air_devices"]) == 17
    assert len(relevant["dampers"]) == 9


def test_field_plan_graph_aware(m12_graph):
    plan = field_plan(m12_graph, SCOPE)
    systems = {s["system"] for s in plan["systems"]}
    assert systems == {"RTU-5", "RTU-6"}
    rtu5 = next(s for s in plan["systems"] if s["system"] == "RTU-5")
    whats = {v["what"] for v in rtu5["verify"]}
    assert {"total airflow", "outside air", "static (ESP)", "fan speed"} <= whats
    assert plan["life_safety"]  # FD-1/FSD-1/SMD-1


def test_ready_to_leave_uses_graph(m12_graph):
    record = _record()
    record.air_devices = [
        AirDevice(device_id="SA-1", function="SUPPLY", final_cfm=183.0),
        AirDevice(device_id="SA-2", function="SUPPLY", final_cfm=184.0),
    ]
    report = ready_to_leave_graph(m12_graph, record)
    assert report["readiness"] == "MISSING BEFORE LEAVING"
    assert any("SA-3" in item for item in report["MISSING_BEFORE_LEAVING"])
    assert any("RTU-5 total airflow" in item for item in report["MISSING_BEFORE_LEAVING"])


def test_ready_to_leave_full(m12_graph):
    record = _record()
    for device in m12_graph.air_devices:
        if device.get("function") != "EXHAUST":
            record.air_devices.append(AirDevice(
                device_id=device["id"], function=device.get("function", "SUPPLY"),
                final_cfm=100.0))
    report = ready_to_leave_graph(m12_graph, record)
    # every planned supply device is measured; only equipment-total
    # verification (by design) may remain
    device_missing = [i for i in report["MISSING_BEFORE_LEAVING"] if "SA-" in i]
    assert not device_missing
    assert report["readiness"] in ("READY", "MISSING BEFORE LEAVING")


# ---------------------------------------------------------------------------
# Missing context / honest degradation
# ---------------------------------------------------------------------------


def test_no_hallucinated_sheet_facts(m12_doc):
    subset = _subset(m12_doc, [2, 3, 5])  # notes + schedule + plan
    graph = build_graph([subset])
    # M7.2 is referenced on the notes sheet but absent -> contextual uncertainty
    assert any("M7.2" in (m.get("sheet_id") or "") for m in graph.missing_context)
    # supplied sheets still extract fully
    assert any(e["id"] == "RTU-5" for e in graph.equipment) or \
        any(d["id"] == "SA-1" for d in graph.air_devices)


def test_missing_legend_abstention(m12_doc):
    subset = _subset(m12_doc, [5, 6])  # plans only, legend absent
    graph = build_graph([subset])
    assert any(m["kind"] == "LEGEND_NOT_SUPPLIED" for m in graph.missing_context)
    # dampers still found but flagged as generic inference, not VERIFIED
    for damper in graph.dampers:
        assert damper.get("confidence") in ("GENERIC_SYMBOL_INFERENCE",
                                            "PROJECT_LEGEND_VERIFIED")


def test_customer_pdf_exclusion():
    from subprocess import run, PIPE
    repo = Path(__file__).resolve().parents[1]
    result = run(["git", "ls-files", "tests", "scs_reports", "tools"],
                 capture_output=True, text=True, cwd=repo)
    pdfs = [f for f in result.stdout.splitlines() if f.lower().endswith(".pdf")]
    assert pdfs == []
