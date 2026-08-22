"""EXACT WORKBOOK AUDIT - canonical sheet identifiers vs current masters.

Asserts the canonical sheet registry (scs_reports.masters) matches the ACTUAL
owner master files in C:\\SCS_DATA\\masters, including exact sheet names
(trailing whitespace and all). Also asserts the Monday-relevant sheets exist.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from scs_reports.masters import MASTER_WORKBOOKS, MasterWorkbook

MASTER_DIR = Path(r"C:\SCS_DATA\masters")
REQUIRED_MASTERS = (
    "Field_Report_Master.xlsm",
    "Test and Balance MASTER TEMPLATE 001.xlsx",
)
# Monday report sheets that MUST exist in Field_Report_Master.xlsm
MONDAY_FIELD_SHEETS = (
    "02_Cover", "03_Executive_Summary", "20_Air_Distribution",
    "22_Duct_Traverse", "19_Static_Pressure_Profile", "30_Deficiencies",
    "05_Instrument_Calibration", "07_Equipment_Register", "31_Photo_Log",
    "32_Final_Closeout",
)


def _sheetnames(filename: str) -> set[str]:
    path = MASTER_DIR / filename
    if not path.exists():
        pytest.skip(f"master not present locally: {path}")
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return set(wb.sheetnames)
    finally:
        wb.close()


def test_primary_masters_present_locally():
    missing = [name for name in REQUIRED_MASTERS if not (MASTER_DIR / name).exists()]
    assert missing == [], f"primary masters missing: {missing}"


def test_registry_sheets_exist_in_field_master():
    workbook = next(w for w in MASTER_WORKBOOKS if w.filename == "Field_Report_Master.xlsm")
    names = _sheetnames(workbook.filename)
    missing = [s.sheet for s in workbook.sheets if s.sheet not in names]
    assert missing == [], f"registry sheets missing from master: {missing}"


def test_registry_sheets_exist_in_tab_master():
    workbook = next(w for w in MASTER_WORKBOOKS if w.filename == "Test and Balance MASTER TEMPLATE 001.xlsx")
    names = _sheetnames(workbook.filename)
    missing = [s.sheet for s in workbook.sheets if s.sheet not in names]
    assert missing == [], f"registry sheets missing from TAB master: {missing}"


def test_monday_sheets_present_in_field_master():
    names = _sheetnames("Field_Report_Master.xlsm")
    missing = [s for s in MONDAY_FIELD_SHEETS if s not in names]
    assert missing == [], f"Monday sheets missing from Field_Report_Master.xlsm: {missing}"


def test_registry_names_unique_and_nonempty():
    """No duplicate or empty sheet names in the registry per workbook."""
    seen: dict[str, set[str]] = {}
    for workbook in MASTER_WORKBOOKS:
        names = [s.sheet for s in workbook.sheets]
        assert all(name.strip() for name in names), (
            f"registry contains empty sheet name in {workbook.filename}"
        )
        assert len(names) == len(set(names)), (
            f"registry duplicate sheet names in {workbook.filename}"
        )
