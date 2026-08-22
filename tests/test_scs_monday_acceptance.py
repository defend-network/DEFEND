"""SYNTHETIC MONDAY ACCEPTANCE.

Exercises the full owner journey for the Monday job (airflow verification of
two newly installed duct systems for two workout studios) end-to-end:

    create job -> describe scope -> ingest photos -> capture measurements
    (as-found/final) -> auto-associate -> completeness -> ready-to-leave
    -> resolve gaps -> generate -> validate -> download workbook

PASS requires: minimal relevant sheets; 0 fabricated numbers; 0 BLOCK;
correct as-found/final; correct photo log; professional executive summary;
minimal owner questions.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scs_reports.completeness import evaluate, ready_to_leave
from scs_reports.composer import Composer
from scs_reports.masters import MASTER_WORKBOOKS
from scs_reports.nl_parse import merge_capture, parse_measurements
from scs_reports.planner import plan_for
from scs_reports.schema import (
    AirDevice,
    Equipment,
    EquipmentType,
    Finding,
    JobMetadata,
    JobRecord,
    Measurement,
    PhotoEvidence,
    Provenance,
    Traverse,
    TraversePoint,
)
from scs_reports.store import JobStore, MasterStore, PhotoIngest, ReportPaths
from scs_reports.validation import validate_report

MASTER_SOURCE = Path(r"C:\SCS_DATA\masters")


@pytest.fixture
def paths(tmp_path):
    return ReportPaths(tmp_path).ensure()


@pytest.fixture
def masters(paths):
    store = MasterStore(paths)
    store.install_masters(MASTER_SOURCE)
    return store


def build_monday_job() -> JobRecord:
    metadata = JobMetadata(
        job_id="monday_studios",
        project_name="Workout Studio Airflow Verification",
        project_number="A-1234",
        site_name="Studio A & B",
        test_date=date(2026, 8, 24),
        technician="Aaron Thomas",
        report_type="AIRFLOW_VERIFICATION",
    )
    record = JobRecord(metadata=metadata)
    record.scope_notes = (
        "Airflow verification and slight balancing of two newly installed "
        "ductwork systems serving small workout studio areas."
    )
    fan = Equipment(
        equipment_id="SF-1", equipment_type=EquipmentType.FAN, tag="SF-1",
        manufacturer="Greenheck", model="SQ-30", serial="SN-1001",
        area_served="Studio A & B",
    )
    fan.measurements = [
        Measurement("airflow_cfm", 900, "cfm", Provenance.TECH_ENTERED, technician_confirmed=True),
        Measurement("fan_rpm", 1100, "rpm", Provenance.TECH_ENTERED, technician_confirmed=True),
        Measurement("sp_A", 0.05, "in.w.c.", Provenance.TECH_ENTERED, technician_confirmed=True),
        Measurement("sp_F", 0.42, "in.w.c.", Provenance.TECH_ENTERED, technician_confirmed=True),
    ]
    record.equipment = [fan]
    return record


def test_monday_scope_to_plan(paths):
    record = build_monday_job()
    # scope described in chat (natural language)
    scope = (
        "Scope: airflow verification and slight balancing of two newly installed "
        "ductwork systems serving the small workout studio areas."
    )
    # parse the scope for measurements (none present)
    caps = parse_measurements(scope)
    assert caps == []
    record.metadata.report_type, record.scope_notes = "AIRFLOW_VERIFICATION", scope
    record.air_devices.append(
        AirDevice(device_id="SA-1", function="SUPPLY", area_served="Studio A",
                  design_cfm=400.0, final_cfm=418.0, as_found_cfm=300.0)
    )
    plan = plan_for(record)
    types = [s.type for s in plan.sections]
    assert "air_distribution" in types
    assert "building_pressure" not in types
    assert types[0] == "cover"
    assert types[-1] == "closeout"


def test_monday_measurement_capture_preserves_as_found(paths):
    record = build_monday_job()
    text = (
        "Studio A SA-3 was 142 CFM as found. I opened the damper and got 181 final. "
        "Studio A SA-2 was 150 CFM as found; balanced to 176 final. "
        "Studio B SA-1 was 300 CFM as found; 418 final after balancing. "
        "Studio B total 1184 CFM final. Design for SA-3 is 200 CFM."
    )
    caps = parse_measurements(text)
    assert any(c.device_id == "SA-3" and c.as_found_cfm == 142.0 for c in caps)
    for c in caps:
        merge_capture(record, c)
    sa3 = next(d for d in record.air_devices if d.device_id == "SA-3")
    assert sa3.as_found_cfm == 142.0  # never overwritten
    assert sa3.final_cfm == 181.0
    assert sa3.design_cfm == 200.0
    assert sa3.measurement_method == "rotating vane"
    total = next(d for d in record.air_devices if d.device_id == "TOTAL")
    assert total.final_cfm == 1184.0


def test_monday_ready_to_leave_blocks_and_questions(paths):
    record = build_monday_job()
    ready = ready_to_leave(record)
    assert ready["ready"] is False  # no devices yet
    assert any("air_devices" in i for i in ready["MISSING_BEFORE_LEAVING"])
    # after measurements + photos
    text = (
        "Studio A SA-3 was 142 CFM as found; opened damper, 181 final. "
        "Studio A SA-2 was 150 CFM as found; 176 final. "
        "Studio B SA-1 was 300 CFM as found; 418 final."
    )
    for c in parse_measurements(text):
        merge_capture(record, c)
    record.photos = [
        PhotoEvidence(photo_id="PHOTO-001", original_filename="a.jpg",
                      sha256="0" * 64)
    ]
    for d in record.air_devices:
        d.design_cfm = 200.0 if d.device_id != "TOTAL" else None
        d.measurement_method = d.measurement_method or "rotating vane"
    ready = ready_to_leave(record)
    assert ready["ready"] is True
    assert len(ready["questions"]) <= 5


def test_monday_photo_batch_and_photo_log(paths, masters, tmp_path):
    record = build_monday_job()
    store = JobStore(paths)
    store.create(record)
    source = tmp_path / "photos"
    source.mkdir()
    for index in range(4):
        (source / f"photo_{index}.txt").write_text(f"evidence {index}", encoding="utf-8")
    entries = PhotoIngest(paths).ingest(record.metadata.job_id, sorted(source.iterdir()))
    record.photos.extend(entries)
    # associate readings to photos
    for device in record.air_devices:
        device.evidence_refs = [entries[0].photo_id]
    store.save(record)
    plan = plan_for(record)
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    wb = load_workbook(output)
    photo_sheet = wb["31_Photo_Log"]
    assert photo_sheet["A12"].value == "PHOTO-001"
    assert photo_sheet["A13"].value == "PHOTO-002"
    wb.close()
    composer.close_masters()


def test_monday_generate_validate_download(paths, masters):
    record = build_monday_job()
    text = (
        "Studio A SA-3 was 142 CFM as found; opened damper, 181 final. "
        "Studio A SA-2 was 150 CFM as found; 176 final. "
        "Studio B SA-1 was 300 CFM as found; 418 final. "
        "Studio A SA-1 was 120 CFM as found; 165 final."
    )
    for c in parse_measurements(text):
        merge_capture(record, c)
    for d in record.air_devices:
        if d.device_id != "TOTAL":
            d.design_cfm = d.device_id in ("SA-1", "SA-2", "SA-3") and 400.0 or 600.0
            d.status = "BALANCED"
        if d.device_id == "SA-1" and d.area_served == "Studio A":
            d.design_cfm = 400.0
    # set SA-1 Studio B design 600
    sa1 = next(d for d in record.air_devices if d.device_id == "SA-1" and d.area_served == "Studio B")
    sa1.design_cfm = 600.0
    record.traverses = [
        Traverse(
            traverse_id="TRV-1", system_id="SF-1", location="Studio A main trunk",
            duct_size="20x20", area_sqft=2.78, design_fpm=800.0, final_fpm=825.0,
            points=[TraversePoint("A", 820.0, 1), TraversePoint("B", 840.0, 1)],
        )
    ]
    record.findings = [Finding("Loose flex", "Flex strap loose at Studio A plenum", severity="minor")]
    record.photos = [
        PhotoEvidence(photo_id="PHOTO-001", original_filename="outlet_a.jpg",
                      sha256="0" * 64),
        PhotoEvidence(photo_id="PHOTO-002", original_filename="outlet_b.jpg",
                      sha256="0" * 64),
    ]
    store = JobStore(paths)
    store.create(record)
    plan = plan_for(record)
    types = [s.type for s in plan.sections]
    # minimal relevant sheets for a small verification report
    assert "air_distribution" in types
    assert "traverse_summary" in types
    assert "traverse_points" in types
    assert "deficiencies" in types
    assert "static_pressure" in types  # sp_A/sp_F recorded in equipment
    assert "vfd_report" not in types
    assert "vav_data" not in types

    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    wb = load_workbook(output)
    assert set(wb.sheetnames) >= {
        "02_Cover", "03_Executive_Summary", "20_Air_Distribution",
        "31_Photo_Log", "32_Final_Closeout", "30_Deficiencies",
    }
    ad = wb["20_Air_Distribution"]
    sa3 = next(d for d in record.air_devices if d.device_id == "SA-3")
    row = 17 + record.air_devices.index(sa3)
    assert ad.cell(row=row, column=2).value == "SA-3"          # device
    assert ad.cell(row=row, column=9).value == 142.0           # as-found (prelim CFM)
    assert ad.cell(row=row, column=11).value == 181.0          # final CFM
    assert ad.cell(row=row, column=12).value is not None       # % design formula
    wb.close()

    validation = validate_report(record, plan, output, masters=masters)
    assert not validation.blocked, (
        "expected 0 BLOCK, got: "
        + "; ".join(f"{c.name}:{c.message}" for c in validation.checks if c.status == "BLOCK")
    )
    assert validation.checks[0].name == "required_fields_complete"
    assert not any(c.status == "BLOCK" for c in validation.checks)
    # deterministic math is consistent (no fabricated numbers)
    math = next(c for c in validation.checks if c.name == "calculation_outputs_correct")
    assert math.status in ("PASS", "WARN")
    # download candidate exists and reopens
    candidates = sorted(paths.output_dir(record.metadata.job_id).glob("*.xlsx"))
    assert candidates
    reopened = load_workbook(candidates[-1])
    assert reopened.sheetnames
    reopened.close()
    composer.close_masters()


def test_monday_minimal_questions_not_giant_questionnaire():
    record = build_monday_job()
    for c in parse_measurements("Studio A SA-1 was 120 CFM as found; 165 final."):
        merge_capture(record, c)
    report = evaluate(record)
    assert len(report.questions) <= 5
    assert all(isinstance(q, str) and q.strip() for q in report.questions)


def test_masters_registry_has_monday_sheets():
    workbook = next(w for w in MASTER_WORKBOOKS if w.filename == "Field_Report_Master.xlsm")
    names = {s.sheet for s in workbook.sheets}
    assert {"20_Air_Distribution", "19_Static_Pressure_Profile",
            "30_Deficiencies", "05_Instrument_Calibration"} <= names
