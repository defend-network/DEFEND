"""SCS Blueprint Vision M1.1 - raster blueprint intelligence tests.

Covers the P45 list for image-based prints: raster detection, rendering,
tiling, OCR geometry, orientation, raster schedule/device/CFM extraction,
OCR/VLM reconciliation + numeric confidence guards, room/system association,
provenance, source crops, corrections, cache/versioning, honest degradation,
raster Plan Chat, pre-engineered report, measured-blank discipline, and
regression guarantees.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from fixtures.blueprint_ground_truth import (
    EXPECTED_DEVICES,
    EXPECTED_EQUIPMENT,
    EXPECTED_SHEETS,
    EXPECTED_SUPPLY_TOTALS,
)
from fixtures.make_blueprint import build_blueprint, build_raster_blueprint
from scs_reports.blueprint_raster import (
    _numeric_status,
    _read_value_crop,
    merge_basis,
    ocr_page_words_pdf,
    page_raster_state,
    page_tiles,
    raster_run,
    run_blueprint,
    source_crop_png,
)
from scs_reports.completeness import ready_to_leave
from scs_reports.corrections import (
    apply_known_correction,
    load_corrections,
    save_correction,
)
from scs_reports.planner import plan_for
from scs_reports.plans import (
    SCS_EXTRACTION_VERSION,
    cached_basis,
    index_pdf,
    sha256_of,
)
from scs_reports.preengineer import build_preengineered_record
from scs_reports.schema import JobMetadata, JobRecord
from scs_reports.store import JobStore, MasterStore, ReportPaths
from scs_reports.validation import validate_report
from scs_reports.vision import DocumentReader, OcrWord, RapidOcrDocumentReader
from tools.blueprint_benchmark import evaluate as benchmark_evaluate

MASTER_SOURCE = Path(r"C:\SCS_DATA\masters")


class NoopDocumentReader(DocumentReader):
    """Honest degradation: no OCR available."""

    def read_page_words(self, image_path: Path):
        return []

    def read_region(self, image_path: Path, crop, prompt):
        return []


@pytest.fixture(scope="module")
def blueprint_pdf(tmp_path_factory) -> Path:
    path = tmp_path_factory.mktemp("bp") / "blueprint_native.pdf"
    build_blueprint(path)
    return path


@pytest.fixture(scope="module")
def raster_pdf(tmp_path_factory) -> Path:
    src = tmp_path_factory.mktemp("bp") / "blueprint_src.pdf"
    build_blueprint(src)
    out = tmp_path_factory.mktemp("bp") / "blueprint_raster.pdf"
    return build_raster_blueprint(src, out)


@pytest.fixture(scope="module")
def raster_basis(raster_pdf, tmp_path_factory):
    cache = tmp_path_factory.mktemp("cache")
    return raster_run(raster_pdf, RapidOcrDocumentReader(),
                      cache_dir=cache)


@pytest.fixture(scope="module")
def raster_conflict_pdf(tmp_path_factory) -> Path:
    src = tmp_path_factory.mktemp("bp") / "conflict_src.pdf"
    build_blueprint(src, supply_overrides={"RTU-5": 1999})
    out = tmp_path_factory.mktemp("bp") / "conflict_raster.pdf"
    return build_raster_blueprint(src, out)


@pytest.fixture
def paths(tmp_path):
    return ReportPaths(tmp_path).ensure()


@pytest.fixture
def masters(paths):
    store = MasterStore(paths)
    store.install_masters(MASTER_SOURCE)
    return store


def _record() -> JobRecord:
    return JobRecord(JobMetadata(
        job_id="raster_monday", project_name="Workout Studio Airflow",
        project_number="A-1234", site_name="Studio A & B",
        test_date=None, technician="AT", report_type="AIRFLOW_VERIFICATION",
    ))


# ---------------------------------------------------------------------------
# Raster detection / rendering / orientation
# ---------------------------------------------------------------------------


def test_raster_detection_flags_image_pages(raster_pdf):
    doc = index_pdf(raster_pdf, tables=False)
    for page in doc.pages:
        assert page_raster_state(page, has_images=True) == "RASTER_PAGE"


def test_native_rich_page_avoids_raster(blueprint_pdf):
    doc = index_pdf(blueprint_pdf, tables=False)
    for page in doc.pages:
        assert page_raster_state(page, has_images=False) == "NATIVE_TEXT_GOOD"


def test_ocr_words_have_geometry(raster_pdf, tmp_path):
    words = ocr_page_words_pdf(raster_pdf, 5, RapidOcrDocumentReader(), 200,
                               tmp_path / "cache")
    assert words
    for word in words[:10]:
        assert word.text and word.x1 > word.x0 and word.bottom > word.top


def test_tile_coordinates_map_to_pdf(raster_pdf):
    tiles = page_tiles(raster_pdf, 5, dpi=200, cols=2, rows=2, overlap_px=40)
    assert len(tiles) == 4
    for tile in tiles:
        x0, y0, x1, y1 = tile.pdf_bbox
        assert x0 < x1 and y0 < y1
        px = tile.pixel_bbox
        # pdf->pixel = * (dpi/72)
        assert abs(x0 * (200 / 72) - px[0]) < 1
        assert abs(y1 * (200 / 72) - px[3]) < 1


def test_overlapping_tiles_deduped_on_merge(raster_basis):
    # two overlapping tile OCR passes must not duplicate device facts
    merged = merge_basis(raster_basis, raster_basis)
    ids = [d["device_id"] for d in merged["instances"]]
    assert len(ids) == len(set(ids))


def test_rotated_sheet_recovers_words(tmp_path_factory, raster_pdf):
    """A 90-degree-rotated page still OCRs (bounded orientation handling)."""
    import fitz
    from PIL import Image

    cache = tmp_path_factory.mktemp("cache")
    rotated_pdf = tmp_path_factory.mktemp("bp") / "rotated.pdf"
    src = fitz.open(str(raster_pdf))
    out = fitz.open()
    page = src.load_page(4)
    pix = page.get_pixmap(matrix=fitz.Matrix(200 / 72, 200 / 72))
    raw_png = tmp_path_factory.mktemp("img") / "page.png"
    pix.save(str(raw_png))
    img = Image.open(raw_png).rotate(90, expand=True)
    rot_png = tmp_path_factory.mktemp("img") / "rot.png"
    img.save(str(rot_png))
    new_page = out.new_page(width=page.rect.height, height=page.rect.width)
    new_page.insert_image(fitz.Rect(0, 0, page.rect.height, page.rect.width),
                          stream=rot_png.read_bytes())
    src.close()
    out.save(str(rotated_pdf))
    out.close()
    from scs_reports.blueprint_raster import _ocr_with_orientation
    words, orientation = _ocr_with_orientation(
        rotated_pdf, 1, RapidOcrDocumentReader(), 200, cache, None)
    assert words, "rotated page should still yield OCR words"
    assert orientation in ("NORMAL", "ROTATED_90", "ROTATED_180", "ROTATED_270")


# ---------------------------------------------------------------------------
# Raster sheet + schedule extraction
# ---------------------------------------------------------------------------


def test_raster_sheet_classification(raster_basis):
    by_page = {c["page"]: c["type"] for c in raster_basis["sheet_classification"]}
    for page, (_sheet, ptype) in EXPECTED_SHEETS.items():
        assert by_page.get(page) == ptype, f"raster page {page}"


def test_raster_schedule_reconstructed(raster_basis):
    assert "SD-1" in raster_basis["schedule_types"]
    assert raster_basis["schedule_types"]["SD-1"]["design_cfm"] == "200"
    equipment = {e["tag"]: e for e in raster_basis["equipment"]}
    assert equipment["RTU-5"]["supply_cfm"] == 1180.0
    assert equipment["RTU-6"]["supply_cfm"] == 1240.0


def test_raster_multi_word_headers(raster_basis):
    assert raster_basis["schedule_types"]["SD-2"]["size"] == "8x8"


# ---------------------------------------------------------------------------
# Raster devices / CFM / size / rooms / totals
# ---------------------------------------------------------------------------


def test_raster_device_cfm_and_size(raster_basis):
    actual = {d["device_id"]: d for d in raster_basis["instances"]}
    for tag, attrs in EXPECTED_DEVICES.items():
        assert tag in actual, tag
        assert actual[tag]["design_cfm"] == attrs["cfm"], f"{tag} CFM"
        assert actual[tag]["size"] == attrs["size"], f"{tag} size"


def test_raster_tag_cfm_spatial_association(raster_basis):
    actual = {d["device_id"]: d for d in raster_basis["instances"]}
    # SA-3 must be 200 (its own callout), never cross-associated with SA-2/SA-4
    assert actual["SA-3"]["design_cfm"] == 200.0
    assert actual["SA-5"]["design_cfm"] == 210.0
    assert actual["SA-11"]["design_cfm"] == 190.0


def test_raster_rooms_and_totals(raster_basis):
    rooms = {r["name"].upper(): r for r in raster_basis["rooms"]}
    assert "WORKOUT STUDIO A" in rooms and "WORKOUT STUDIO B" in rooms
    totals = {t["scope"]: t for t in raster_basis["design_totals"]
              if t["function"] == "SUPPLY"}
    assert totals["WORKOUT STUDIO A"]["design_total_cfm"] == 1180.0
    assert totals["WORKOUT STUDIO B"]["design_total_cfm"] == 1240.0


def test_raster_provenance_complete(raster_basis):
    for device in raster_basis["instances"]:
        source = device["source"]
        assert source.get("page")
        assert source.get("bbox")
        assert source.get("extraction_method") in ("OCR_PLAN", "OCR_SCHEDULE")
        if device.get("design_cfm") is not None:
            assert device.get("numeric_status") in ("VERIFIED", "HIGH")


def test_raster_system_association(raster_basis):
    associations = {
        a["system_id"]: a for a in raster_basis.get("system_associations", [])
    }
    assert "RTU-5" in associations
    assert associations["RTU-5"]["confidence"] == "HIGH"
    assert "SA-1" in associations["RTU-5"]["devices_served"]
    assert any(e["kind"] in ("SCHEDULE_REMARKS", "SCHEDULE_SUPPLY_MATCH")
               for e in associations["RTU-5"]["evidence"])


# ---------------------------------------------------------------------------
# Numeric confidence guards
# ---------------------------------------------------------------------------


def test_numeric_status_agreement_verified():
    assert _numeric_status(185.0, 0.7, [("185", 0.9)]) == "VERIFIED"


def test_numeric_status_disagreement_conflict():
    assert _numeric_status(185.0, 0.7, [("165", 0.9)]) == "CONFLICT"


def test_numeric_status_low_confidence_review():
    assert _numeric_status(185.0, 0.3, []) == "REVIEW_REQUIRED"


def test_low_confidence_never_autofills(raster_basis):
    record = build_preengineered_record(_record(), raster_basis)
    for device in record.air_devices:
        assert device.design_cfm is not None  # all VERIFIED/HIGH auto-populated
    # a REVIEW_REQUIRED device must abstain
    basis = {
        "equipment": [], "rooms": [], "design_totals": [], "conflicts": [],
        "instances": [{
            "device_id": "SA-99", "room": None, "design_cfm": None, "size": None,
            "schedule_type": None, "confidence": "REVIEW_REQUIRED",
            "numeric_status": "REVIEW_REQUIRED",
            "source": {"sheet": "M9", "page": 9, "bbox": (0, 0, 1, 1),
                       "extraction_method": "OCR_PLAN"},
        }],
    }
    record2 = build_preengineered_record(_record(), basis)
    dev = next(d for d in record2.air_devices if d.device_id == "SA-99")
    assert dev.design_cfm is None


def test_raster_conflict_surfaces(raster_conflict_pdf, tmp_path_factory):
    cache = tmp_path_factory.mktemp("cache")
    basis = raster_run(raster_conflict_pdf, RapidOcrDocumentReader(),
                       cache_dir=cache)
    assert basis["conflicts"]
    assert any(c["kind"] == "DESIGN_DOCUMENT_CONFLICT"
               for c in basis["conflicts"])


# ---------------------------------------------------------------------------
# Source crops / corrections / cache
# ---------------------------------------------------------------------------


def test_source_crop_generated(raster_pdf, tmp_path_factory):
    cache = tmp_path_factory.mktemp("cache")
    png = source_crop_png(raster_pdf, 5, (150, 430, 260, 490), cache)
    assert png.startswith(b"\x89PNG")


def test_correction_preserves_original(tmp_path):
    store = tmp_path / "corrections.json"
    save_correction(store, sha256="abc123", device="SA-3",
                    original_value=165, corrected_value=185, sheet="M2.1",
                    corrected_by="owner", reason="owner confirmed")
    records = load_corrections(store)
    assert records[list(records)[0]]["original_value"] == 165
    assert records[list(records)[0]]["corrected_value"] == 185
    assert apply_known_correction(store, "abc123", "SA-3", "M2.1") == 185


def test_cache_reuse_and_version_invalidation(blueprint_pdf, tmp_path):
    cache_dir = tmp_path / "cache"
    first = cached_basis(blueprint_pdf, cache_dir)
    second = cached_basis(blueprint_pdf, cache_dir)
    assert first["_cache_hit"] is False
    assert second["_cache_hit"] is True
    versioned = cache_dir / f"{SCS_EXTRACTION_VERSION}_{sha256_of(blueprint_pdf)}.json"
    assert versioned.exists()


def test_raster_cache_reused(raster_pdf, tmp_path_factory):
    cache = tmp_path_factory.mktemp("cache")
    basis1 = raster_run(raster_pdf, RapidOcrDocumentReader(), cache_dir=cache)
    basis2 = raster_run(raster_pdf, RapidOcrDocumentReader(), cache_dir=cache)
    assert [d["device_id"] for d in basis1["instances"]] == \
        [d["device_id"] for d in basis2["instances"]]
    # render cache avoids re-rendering
    renders = list((cache).glob("*.png"))
    assert renders


# ---------------------------------------------------------------------------
# Honest degradation / merged pipeline / chat / report
# ---------------------------------------------------------------------------


def test_ocr_unavailable_degrades_honestly(raster_pdf, tmp_path_factory):
    cache = tmp_path_factory.mktemp("cache")
    basis = run_blueprint(raster_pdf, NoopDocumentReader(), cache_dir=cache)
    # no crash; native cover/notes may still be absent but pipeline returns
    assert isinstance(basis["instances"], list)
    assert isinstance(basis["sheet_classification"], list)


def test_merged_native_plus_raster(raster_pdf, tmp_path_factory):
    cache = tmp_path_factory.mktemp("cache")
    basis = run_blueprint(raster_pdf, RapidOcrDocumentReader(), cache_dir=cache)
    assert basis["extraction"]["mode"] == "RASTER"
    assert len(basis["instances"]) == len(EXPECTED_DEVICES)


def test_raster_plan_chat_cites_sheet(raster_basis):
    from tools.job_copilot import answer_plan_question
    answer = answer_plan_question("What is SA-6 supposed to be?", raster_basis)
    assert "210" in answer["answer"]
    assert answer["source"]["sheet"] == "M3.1"
    assert answer["source"]["page"] == 5
    total = answer_plan_question("What is the design total for Studio A?",
                                 raster_basis)
    assert "1180" in total["answer"]


def test_report_design_rows_and_measured_blank(paths, masters, raster_basis):
    record = build_preengineered_record(_record(), raster_basis)
    store = JobStore(paths)
    store.create(record)
    plan = plan_for(record)
    from scs_reports.composer import Composer
    composer = Composer(paths, store)
    output = composer.compose(record, plan)
    wb = load_workbook(output)
    ad = wb["20_Air_Distribution"]
    assert ad.cell(row=17, column=2).value == "SA-1"
    assert ad.cell(row=17, column=7).value == 180.0      # design CFM
    assert ad.cell(row=17, column=9).value is None        # as-found blank
    assert ad.cell(row=17, column=11).value is None       # final blank
    wb.close()
    composer.close_masters()


def test_ready_to_leave_catches_omitted(raster_basis):
    record = build_preengineered_record(_record(), raster_basis)
    report = ready_to_leave(record)
    assert report["ready"] is False
    assert any("planned_devices_measured" in i
               for i in report["MISSING_BEFORE_LEAVING"])


def test_raster_benchmark_metrics(raster_basis):
    ground_truth = {
        "EXPECTED_SHEETS": EXPECTED_SHEETS,
        "EXPECTED_EQUIPMENT": EXPECTED_EQUIPMENT,
        "EXPECTED_DEVICES": EXPECTED_DEVICES,
        "EXPECTED_SUPPLY_TOTALS": EXPECTED_SUPPLY_TOTALS,
    }
    metrics = benchmark_evaluate(raster_basis, ground_truth)
    assert metrics["CFM_EXACT_MATCH"] == 1.0
    assert metrics["FALSE_CFM_RATE"] == 0.0
    assert metrics["UNSUPPORTED_AUTOFILL_RATE"] == 0.0
    assert metrics["SIZE_EXACT_MATCH"] == 1.0
    assert metrics["ROOM_ASSOCIATION_ACCURACY"] == 1.0
    assert metrics["DESIGN_TOTAL_EXACT_MATCH"] == 1.0
    assert metrics["DEVICE_TAG_RECALL"] == 1.0
